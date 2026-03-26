# views.py
from io import BytesIO
import io
import os
import logging
from django.db.models import Q
from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.template.loader import render_to_string
from django.urls import reverse
from django.http import HttpResponse
#from weasyprint import HTML, CSS
from xhtml2pdf import pisa
from decimal import Decimal
import tempfile
from django.conf import settings
from django.core.mail import EmailMessage
from datetime import datetime
from django.template.loader import get_template

import csv
from django.views import View
from FE.models import FacturaElectronica
from INVENTARIO.models import Compra
from django.db.models import Sum, F, Value, DecimalField
from .models import (
    CuentaContable, AsientoContable, LineaAsiento,
    CuentaPorCobrar, PagoCobro,
    CuentaPorPagar, PagoPagar,
)
from FE.models import Receptor_fe
from INVENTARIO.models import Proveedor

logger = logging.getLogger(__name__)

###################################################################################################################


# ---------------------------------
# ANEXOS DE HACIENDA
# ---------------------------------

# Helpers de formato
def padr(s: str, width: int, char: str = " ") -> str:
    return (s or "").ljust(width, char)[:width]

def padl(s: str, width: int, char: str = " ") -> str:
    return (s or "").rjust(width, char)[-width:]

def format_moneda(v) -> str:
    # dos decimales sin separador de miles
    return f"{v:,.2f}".replace(",", "") if v is not None else "0.00"

# Anexo de Ventas a Consumidor Final
class AnexoConsumidorFinalCSV(View):
    """
    CSV Anexo Consumidor Final para facturas de tipo_dte.codigo = '01'.
    Agrega BOM para que Excel lea correctamente los acentos.
    """

    def get(self, request, *args, **kwargs):
        # Solo facturas cuyo tipo_dte.codigo sea '01'
        qs = FacturaElectronica.objects.filter(
            tipo_dte__codigo='01'
        ).order_by('fecha_emision')

        logger.info(f"[Anexo CSV] facturas tipo_dte=01: {qs.count()}")

        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = \
            'attachment; filename="anexo_consumidor_final.csv"'
        response.write('\ufeff')  # BOM para Excel

        writer = csv.writer(response, delimiter=';')

        # Cabecera
        writer.writerow([
            'FECHA DE EMISIÓN',
            'CLASE DE DOCUMENTO',
            'TIPO DE DOCUMENTO',
            'NÚMERO DE RESOLUCIÓN',
            'SERIE DEL DOCUMENTO',
            'NUMERO DE CONTROL INTERNO DEL',
            'NUMERO DE CONTROL INTERNO AL',
            'NÚMERO DE DOCUMENTO (DEL)',
            'NÚMERO DE DOCUMENTO (AL)',
            'NÚMERO DE MAQUINA REGISTRADORA',
            'VENTAS EXENTAS',
            'VENTAS INTERNAS EXENTAS NO SUJETAS A PROPORCIONALIDAD',
            'VENTAS NO SUJETAS',
            'VENTAS GRAVADAS LOCALES',
            'EXPORTACIONES DENTRO DEL ÁREA DE CENTROAMÉRICA',
            'EXPORTACIONES FUERA DEL ÁREA DE CENTROAMÉRICA',
            'EXPORTACIONES DE SERVICIO',
            'VENTAS A ZONAS FRANCAS Y DPA (TASA CERO)',
            'VENTAS A CUENTA DE TERCEROS NO DOMICILIADOS',
            'TOTAL DE VENTAS',
            'TIPO DE OPERACIÓN (RENTA)',
            'TIPO DE INGRESO (RENTA)',
            'NÚMERO DEL ANEXO',
        ])

        for f in qs:
            fecha = f.fecha_emision.strftime('%d/%m/%Y')
            clase = '4'
            # Usamos la descripción que viene de la BD para el tipo 01
            tipo = f"{f.tipo_dte.codigo}"
            num_resol = f.numero_control or ''
            serie = f.codigo_generacion.hex.upper() if f.codigo_generacion else ''

            # Campos internos y máquina vacíos (si no aplica)
            ctrl_del = ctrl_al = doc_del = doc_al = num_reg = ''

            # Montos (asegurar no nulos)
            v_exentas = f.total_exentas or 0
            v_no_sujetas = f.total_no_sujetas or 0
            v_gravadas = f.total_gravadas or 0
            total_ventas = f.total_pagar or 0

            # Operación / ingreso
            tipo_oper = (
                f"{f.condicion_operacion.codigo}"
                if f.condicion_operacion else ''
            )
            tipo_ing = ''  # ajustar si lo calculas distinto

            num_anexo = '2'

            writer.writerow([
                fecha,
                clase,
                tipo,
                num_resol,
                serie,
                ctrl_del,
                ctrl_al,
                doc_del,
                doc_al,
                num_reg,
                f"{v_exentas:.2f}",
                '',  # ventas internas exentas no sujetas
                f"{v_no_sujetas:.2f}",
                f"{v_gravadas:.2f}",
                '0.00', '0.00', '0.00', '0.00', '0.00',
                f"{total_ventas:.2f}",
                tipo_oper,
                tipo_ing,
                num_anexo,
            ])

        return response

# Anexo de Ventas a Contribuyentes
class AnexoContribuyentesCSV(View):
    """
    Genera el CSV ANEXO CONTRIBUYENTES para FacturaElectronica tipo '01'.
    """

    def get(self, request, *args, **kwargs):
        qs = FacturaElectronica.objects.filter(
            tipo_dte__codigo="01"
        ).order_by("fecha_emision")

        logger.info(f"[Anexo Contribuyentes] facturas tipo 01: {qs.count()}")

        resp = HttpResponse(content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = 'attachment; filename="anexo_contribuyentes.csv"'
        resp.write("\ufeff")  # BOM para Excel

        writer = csv.writer(resp, delimiter=";")

        # Cabecera
        writer.writerow([
            "FECHA DE EMISIÓN",
            "CLASE DE DOCUMENTO",
            "TIPO DE DOCUMENTO",
            "NÚMERO DE RESOLUCIÓN",
            "SERIE DEL DOCUMENTO",
            "NÚMERO DE DOCUMENTO",
            "NÚMERO DE CONTROL INTERNO",
            "NIT PROVEEDOR/CLIENTE",
            "NOMBRE RAZÓN SOCIAL",
            "VENTAS EXENTAS",
            "VENTAS NO SUJETAS",
            "VENTAS GRAVADAS LOCALES",
            "DÉBITO FISCAL",
            "VENTAS A TERCEROS NO DOMICILIADOS",
            "DÉBITO TERCEROS",
            "TOTAL DE VENTAS",
            "DUI",
            "TIPO DE OPERACIÓN",
            "TIPO DE INGRESO",
            "NÚMERO DEL ANEXO",
        ])

        for f in qs:
            sFECHADOC    = f.fecha_emision.strftime("%d/%m/%Y")
            sCLASEDOC    = padr("4", 1)   # fijo para DTE
            sTIPODOC     = padr(f.tipo_dte.codigo, 2)
            sNUMRESOL    = padl(f.numero_control, 100)
            sNUMSERIE    = padr(f.codigo_generacion.hex.upper(), 100)
            sNUMDOC      = padr(f.dtereceptor.num_documento or "", 100)
            sNUMINTERNO  = padr("", 100)   # no tiene interno
            sNITPROV     = padl(f.dtereceptor.num_documento or "", 14)
            sNOMBRE      = padr(f.dtereceptor.nombre, 100)

            sVEXENTAS      = padl(format_moneda(f.total_exentas or 0), 11)
            sVNOSUJETAS    = padl(format_moneda(f.total_no_sujetas or 0), 11)
            sVGRAVADAS     = padl(format_moneda(f.total_gravadas or 0), 11)
            sDEBITOFISCAL  = padl(format_moneda(f.iva_retenido or 0), 11)
            sVTERCEROS     = padl(format_moneda(0), 11)
            sDEBITOTERC    = padl(format_moneda(0), 11)
            sTOTALVENTAS   = padl(format_moneda(f.total_pagar or 0), 11)

            sDUI           = padl("", 9)
            sTIPOOPER      = padr(f.condicion_operacion.codigo if f.condicion_operacion else "", 2)
            sTIPOINGRESO   = padr("", 2)
            sNUMANEXO      = padr("1", 1)

            writer.writerow([
                sFECHADOC,
                sCLASEDOC,
                sTIPODOC,
                sNUMRESOL,
                sNUMSERIE,
                sNUMDOC,
                sNUMINTERNO,
                sNITPROV,
                sNOMBRE,
                sVEXENTAS,
                sVNOSUJETAS,
                sVGRAVADAS,
                sDEBITOFISCAL,
                sVTERCEROS,
                sDEBITOTERC,
                sTOTALVENTAS,
                sDUI,
                sTIPOOPER,
                sTIPOINGRESO,
                sNUMANEXO,
            ])

        return resp
    
# Anexo de Compras
class AnexoComprasCSV(View):
    """
    Genera el CSV ANEXO DE COMPRAS (anexo 3).
    """
    def get(self, request, *args, **kwargs):
        qs = Compra.objects.all().order_by('fecha')
        logger.info(f"[Anexo Compras] compras: {qs.count()}")

        resp = HttpResponse(content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = 'attachment; filename="anexo_compras.csv"'
        resp.write('\ufeff')  # BOM para Excel

        writer = csv.writer(resp, delimiter=';')

        # Cabecera:
        writer.writerow([
            'FECHA DE EMISIÓN',
            'CLASE DE DOCUMENTO',
            'TIPO DE DOCUMENTO',
            'NÚMERO DEL DOCUMENTO',
            'NIT PROVEEDOR',
            'NOMBRE PROVEEDOR',
            'COMPRAS INTERNAS EXENTAS',
            'INTERNACIONES EXENTAS Y/O NO SUJETAS',
            'IMPORTACIONES EXENTAS Y/O NO SUJETAS',
            'COMPRAS INTERNAS GRAVADAS',
            'INTERNACIONES GRAVADAS BIENES',
            'IMPORTACIONES GRAVADAS BIENES',
            'IMPORTACIONES GRAVADAS SERVICIOS',
            'CRÉDITO FISCAL',
            'TOTAL DE COMPRAS',
            'DUI',
            'TIPO DE OPERACIÓN',
            'CLASIFICACIÓN',
            'SECTOR',
            'TIPO GASTO/COSTO',
            'NÚMERO DEL ANEXO',
        ])

        for compra in qs:
            # 1) Fechas y documentos: (añade los campos a tu modelo si no existen)
            sFECHADOC    = compra.fecha.strftime('%d/%m/%Y')
            sCLASEDOC    = padr('4', 1)               # fijo DTE
            sTIPODOC     = padr(getattr(compra, 'tipo_documento', ''), 2)
            sNUMDOC      = padr(getattr(compra, 'numero_documento', ''), 100)

            # 2) Proveedor:
            sNITPROV     = padl(compra.proveedor.ruc_nit, 14)
            sNOMBRE      = padr(compra.proveedor.nombre, 100)

            # 3) Montos (aquí debes definir la lógica de cálculo según tus categorías):
            detalles = compra.detalles.all()
            # Ejemplo: todo como ‘gravado’ salvo producto.precio_iva==False
            exentas_int = sum(d.subtotal for d in detalles if not d.producto.precio_iva)
            gravadas_int= sum(d.subtotal for d in detalles if d.producto.precio_iva)
            # Para los demás campos (importaciones, internaciones, etc.) necesitas
            # un flag en DetalleCompra o Producto para diferenciarlos; por ahora cero:
            sCOMPRASEXENTAS     = padl(format_moneda(exentas_int), 11)
            sINTEREXENTAS       = padl('0.00', 11)
            sIMPOREXENTAS       = padl('0.00', 11)
            sCOMPRASGRAVADAS    = padl(format_moneda(gravadas_int), 11)
            sINTERGRAVBIENES    = padl('0.00', 11)
            sIMPORGRAVBIENES    = padl('0.00', 11)
            sIMPORGRAVSERV      = padl('0.00', 11)
            # Crédito fiscal = suma de IVA en detalles (si lo calculas)
            credito_fiscal      = sum(
                (d.subtotal * Decimal('0.13')).quantize(Decimal('0.01'))
                for d in detalles if d.producto.precio_iva
            )
            sCREDITOFISCAL      = padl(format_moneda(credito_fiscal), 11)
            sTOTALCOMPRAS       = padl(format_moneda(compra.total), 11)

            # 4) Otros campos que el macro usa: añade a tu modelo si no existen
            sDUI                = padl('', 9)
            sTIPOOPERACION      = padr(getattr(compra, 'tipo_operacion', ''), 1)
            sCLASIFICACION      = padr(getattr(compra, 'clasificacion', ''), 1)
            sSECTOR             = padr(getattr(compra, 'sector', ''), 1)
            sTIPOCOSTOGASTO     = padr(getattr(compra, 'tipo_costo_gasto', ''), 1)

            sNUMANEXO           = padr('3', 1)

            writer.writerow([
                sFECHADOC,
                sCLASEDOC,
                sTIPODOC,
                sNUMDOC,
                sNITPROV,
                sNOMBRE,
                sCOMPRASEXENTAS,
                sINTEREXENTAS,
                sIMPOREXENTAS,
                sCOMPRASGRAVADAS,
                sINTERGRAVBIENES,
                sIMPORGRAVBIENES,
                sIMPORGRAVSERV,
                sCREDITOFISCAL,
                sTOTALCOMPRAS,
                sDUI,
                sTIPOOPERACION,
                sCLASIFICACION,
                sSECTOR,
                sTIPOCOSTOGASTO,
                sNUMANEXO,
            ])

        return resp


# ─────────────────────────────────────────────────────────────────────────────
# PLAN DE CUENTAS
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def cuentas_lista(request):
    q    = request.GET.get('q', '').strip()
    tipo = request.GET.get('tipo', '').strip()

    qs = CuentaContable.objects.select_related('cuenta_padre').all()
    if q:
        qs = qs.filter(Q(codigo__icontains=q) | Q(nombre__icontains=q))
    if tipo:
        qs = qs.filter(tipo=tipo)

    paginator = Paginator(qs, 25)
    page      = request.GET.get('page')
    cuentas   = paginator.get_page(page)

    return render(request, 'contabilidad/cuentas/lista.html', {
        'cuentas': cuentas,
        'q':       q,
        'f_tipo':  tipo,
        'tipos':   CuentaContable.TIPO_CHOICES,
    })


@login_required
def cuentas_crear(request):
    if request.method == 'POST':
        try:
            cuenta = _cuenta_desde_post(request.POST)
            cuenta.save()
            messages.success(request, f'Cuenta {cuenta.codigo} creada correctamente.')
            return redirect('cont-cuentas-lista')
        except Exception as e:
            messages.error(request, str(e))

    return render(request, 'contabilidad/cuentas/form.html', {
        'titulo':      'Nueva cuenta',
        'tipos':       CuentaContable.TIPO_CHOICES,
        'naturalezas': CuentaContable.NATURALEZA_CHOICES,
        'niveles':     CuentaContable.NIVEL_CHOICES,
        'padres':      CuentaContable.objects.filter(nivel='PADRE', activa=True).order_by('codigo'),
    })


@login_required
def cuentas_editar(request, pk):
    cuenta = get_object_or_404(CuentaContable, pk=pk)

    if request.method == 'POST':
        try:
            _cuenta_desde_post(request.POST, cuenta)
            cuenta.save()
            messages.success(request, f'Cuenta {cuenta.codigo} actualizada.')
            return redirect('cont-cuentas-lista')
        except Exception as e:
            messages.error(request, str(e))

    return render(request, 'contabilidad/cuentas/form.html', {
        'titulo':      'Editar cuenta',
        'cuenta':      cuenta,
        'tipos':       CuentaContable.TIPO_CHOICES,
        'naturalezas': CuentaContable.NATURALEZA_CHOICES,
        'niveles':     CuentaContable.NIVEL_CHOICES,
        'padres':      CuentaContable.objects.filter(nivel='PADRE', activa=True).exclude(pk=pk).order_by('codigo'),
    })


@login_required
def cuentas_plantilla(request):
    """Descarga plantilla Excel con plan de cuentas base de El Salvador."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plan de Cuentas"

    headers = ["Código", "Nombre", "Tipo", "Naturaleza", "Nivel", "Código Padre"]
    header_fill = PatternFill("solid", fgColor="0D47A1")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Plan de cuentas base para El Salvador
    cuentas = [
        # ACTIVOS
        ("1",      "ACTIVO",                              "ACTIVO",  "DEUDORA",   "PADRE", ""),
        ("11",     "ACTIVO CORRIENTE",                    "ACTIVO",  "DEUDORA",   "PADRE", "1"),
        ("1101",   "Caja General",                        "ACTIVO",  "DEUDORA",   "DETALLE", "11"),
        ("1102",   "Bancos",                              "ACTIVO",  "DEUDORA",   "DETALLE", "11"),
        ("110201", "Banco Agrícola",                      "ACTIVO",  "DEUDORA",   "DETALLE", "1102"),
        ("110202", "Banco Cuscatlán",                     "ACTIVO",  "DEUDORA",   "DETALLE", "1102"),
        ("1103",   "Cuentas por Cobrar Comerciales",      "ACTIVO",  "DEUDORA",   "DETALLE", "11"),
        ("1104",   "Inventario de Mercaderías",           "ACTIVO",  "DEUDORA",   "DETALLE", "11"),
        ("1105",   "IVA Crédito Fiscal",                  "ACTIVO",  "DEUDORA",   "DETALLE", "11"),
        ("1106",   "Pagos Anticipados",                   "ACTIVO",  "DEUDORA",   "DETALLE", "11"),
        ("1107",   "Deudores Varios",                     "ACTIVO",  "DEUDORA",   "DETALLE", "11"),
        ("12",     "ACTIVO NO CORRIENTE",                 "ACTIVO",  "DEUDORA",   "PADRE", "1"),
        ("1201",   "Propiedad, Planta y Equipo",          "ACTIVO",  "DEUDORA",   "PADRE", "12"),
        ("120101", "Terrenos",                            "ACTIVO",  "DEUDORA",   "DETALLE", "1201"),
        ("120102", "Edificios",                           "ACTIVO",  "DEUDORA",   "DETALLE", "1201"),
        ("120103", "Mobiliario y Equipo de Oficina",      "ACTIVO",  "DEUDORA",   "DETALLE", "1201"),
        ("120104", "Equipo de Transporte",                "ACTIVO",  "DEUDORA",   "DETALLE", "1201"),
        ("120105", "Equipo de Cómputo",                   "ACTIVO",  "DEUDORA",   "DETALLE", "1201"),
        ("1202",   "Depreciación Acumulada",              "ACTIVO",  "ACREEDORA", "PADRE", "12"),
        ("120201", "Deprec. Edificios",                   "ACTIVO",  "ACREEDORA", "DETALLE", "1202"),
        ("120202", "Deprec. Mobiliario y Equipo",         "ACTIVO",  "ACREEDORA", "DETALLE", "1202"),
        ("120203", "Deprec. Equipo de Transporte",        "ACTIVO",  "ACREEDORA", "DETALLE", "1202"),
        ("120204", "Deprec. Equipo de Cómputo",           "ACTIVO",  "ACREEDORA", "DETALLE", "1202"),
        # PASIVOS
        ("2",      "PASIVO",                              "PASIVO",  "ACREEDORA", "PADRE", ""),
        ("21",     "PASIVO CORRIENTE",                    "PASIVO",  "ACREEDORA", "PADRE", "2"),
        ("2101",   "Cuentas por Pagar Comerciales",       "PASIVO",  "ACREEDORA", "DETALLE", "21"),
        ("2102",   "Documentos por Pagar",                "PASIVO",  "ACREEDORA", "DETALLE", "21"),
        ("2103",   "Préstamos Bancarios a Corto Plazo",   "PASIVO",  "ACREEDORA", "DETALLE", "21"),
        ("2104",   "IVA Débito Fiscal",                   "PASIVO",  "ACREEDORA", "DETALLE", "21"),
        ("2105",   "Retenciones por Pagar",               "PASIVO",  "ACREEDORA", "DETALLE", "21"),
        ("2106",   "Acreedores Varios",                   "PASIVO",  "ACREEDORA", "DETALLE", "21"),
        ("2107",   "Impuestos por Pagar",                 "PASIVO",  "ACREEDORA", "DETALLE", "21"),
        ("2108",   "Sueldos y Salarios por Pagar",        "PASIVO",  "ACREEDORA", "DETALLE", "21"),
        ("22",     "PASIVO NO CORRIENTE",                 "PASIVO",  "ACREEDORA", "PADRE", "2"),
        ("2201",   "Préstamos Bancarios a Largo Plazo",   "PASIVO",  "ACREEDORA", "DETALLE", "22"),
        ("2202",   "Hipotecas por Pagar",                 "PASIVO",  "ACREEDORA", "DETALLE", "22"),
        # CAPITAL
        ("3",      "PATRIMONIO",                          "CAPITAL", "ACREEDORA", "PADRE", ""),
        ("31",     "CAPITAL CONTABLE",                    "CAPITAL", "ACREEDORA", "PADRE", "3"),
        ("3101",   "Capital Social",                      "CAPITAL", "ACREEDORA", "DETALLE", "31"),
        ("3102",   "Reserva Legal",                       "CAPITAL", "ACREEDORA", "DETALLE", "31"),
        ("3103",   "Utilidades Acumuladas",               "CAPITAL", "ACREEDORA", "DETALLE", "31"),
        ("3104",   "Utilidad del Ejercicio",              "CAPITAL", "ACREEDORA", "DETALLE", "31"),
        ("3105",   "Pérdida del Ejercicio",               "CAPITAL", "DEUDORA",   "DETALLE", "31"),
        # INGRESOS
        ("4",      "INGRESOS",                            "INGRESO", "ACREEDORA", "PADRE", ""),
        ("41",     "INGRESOS DE OPERACIÓN",               "INGRESO", "ACREEDORA", "PADRE", "4"),
        ("4101",   "Ventas",                              "INGRESO", "ACREEDORA", "DETALLE", "41"),
        ("4102",   "Devoluciones y Rebajas sobre Ventas", "INGRESO", "DEUDORA",   "DETALLE", "41"),
        ("4103",   "Descuentos sobre Ventas",             "INGRESO", "DEUDORA",   "DETALLE", "41"),
        ("42",     "OTROS INGRESOS",                      "INGRESO", "ACREEDORA", "PADRE", "4"),
        ("4201",   "Ingresos Financieros",                "INGRESO", "ACREEDORA", "DETALLE", "42"),
        ("4202",   "Otros Ingresos no Operacionales",     "INGRESO", "ACREEDORA", "DETALLE", "42"),
        # GASTOS / COSTOS
        ("5",      "COSTOS Y GASTOS",                     "GASTO",   "DEUDORA",   "PADRE", ""),
        ("51",     "COSTO DE VENTAS",                     "GASTO",   "DEUDORA",   "PADRE", "5"),
        ("5101",   "Costo de Ventas",                     "GASTO",   "DEUDORA",   "DETALLE", "51"),
        ("5102",   "Compras",                             "GASTO",   "DEUDORA",   "DETALLE", "51"),
        ("5103",   "Devoluciones y Rebajas sobre Compras","GASTO",   "ACREEDORA", "DETALLE", "51"),
        ("52",     "GASTOS DE OPERACIÓN",                 "GASTO",   "DEUDORA",   "PADRE", "5"),
        ("5201",   "Gastos de Venta",                     "GASTO",   "DEUDORA",   "PADRE", "52"),
        ("520101", "Sueldos y Salarios (Ventas)",         "GASTO",   "DEUDORA",   "DETALLE", "5201"),
        ("520102", "Comisiones sobre Ventas",             "GASTO",   "DEUDORA",   "DETALLE", "5201"),
        ("520103", "Publicidad y Propaganda",             "GASTO",   "DEUDORA",   "DETALLE", "5201"),
        ("520104", "Transporte y Fletes",                 "GASTO",   "DEUDORA",   "DETALLE", "5201"),
        ("520105", "Depreciación (Ventas)",               "GASTO",   "DEUDORA",   "DETALLE", "5201"),
        ("5202",   "Gastos de Administración",            "GASTO",   "DEUDORA",   "PADRE", "52"),
        ("520201", "Sueldos y Salarios (Admin)",          "GASTO",   "DEUDORA",   "DETALLE", "5202"),
        ("520202", "Alquileres",                          "GASTO",   "DEUDORA",   "DETALLE", "5202"),
        ("520203", "Servicios Básicos",                   "GASTO",   "DEUDORA",   "DETALLE", "5202"),
        ("520204", "Papelería y Útiles",                  "GASTO",   "DEUDORA",   "DETALLE", "5202"),
        ("520205", "Honorarios Profesionales",            "GASTO",   "DEUDORA",   "DETALLE", "5202"),
        ("520206", "Seguros",                             "GASTO",   "DEUDORA",   "DETALLE", "5202"),
        ("520207", "Mantenimiento y Reparaciones",        "GASTO",   "DEUDORA",   "DETALLE", "5202"),
        ("520208", "Depreciación (Admin)",                "GASTO",   "DEUDORA",   "DETALLE", "5202"),
        ("520209", "Gastos Diversos",                     "GASTO",   "DEUDORA",   "DETALLE", "5202"),
        ("53",     "GASTOS FINANCIEROS",                  "GASTO",   "DEUDORA",   "PADRE", "5"),
        ("5301",   "Intereses Bancarios",                 "GASTO",   "DEUDORA",   "DETALLE", "53"),
        ("5302",   "Comisiones Bancarias",                "GASTO",   "DEUDORA",   "DETALLE", "53"),
        ("5303",   "Diferencial Cambiario",               "GASTO",   "DEUDORA",   "DETALLE", "53"),
    ]

    for row_idx, (cod, nom, tipo, nat, nivel, padre) in enumerate(cuentas, 2):
        ws.cell(row=row_idx, column=1, value=cod)
        ws.cell(row=row_idx, column=2, value=nom)
        ws.cell(row=row_idx, column=3, value=tipo)
        ws.cell(row=row_idx, column=4, value=nat)
        ws.cell(row=row_idx, column=5, value=nivel)
        ws.cell(row=row_idx, column=6, value=padre)

    # Ajustar anchos
    widths = [14, 45, 12, 14, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Hoja de instrucciones
    wi = wb.create_sheet("Instrucciones")
    instrucciones = [
        "INSTRUCCIONES PARA IMPORTAR PLAN DE CUENTAS",
        "",
        "1. Edite la hoja 'Plan de Cuentas' con sus cuentas contables.",
        "2. Mantenga los encabezados exactamente como están (fila 1).",
        "3. Columnas:",
        "   - Código: Código único de la cuenta (ej: 1101, 110201)",
        "   - Nombre: Nombre descriptivo de la cuenta",
        "   - Tipo: ACTIVO, PASIVO, CAPITAL, INGRESO o GASTO",
        "   - Naturaleza: DEUDORA o ACREEDORA",
        "   - Nivel: PADRE (agrupadora) o DETALLE (recibe movimientos)",
        "   - Código Padre: Código de la cuenta padre (vacío si es raíz)",
        "",
        "4. Las cuentas PADRE no reciben movimientos, solo agrupan.",
        "5. Primero se crean las cuentas padre, luego las de detalle.",
        "6. Al importar, las cuentas existentes (mismo código) se omiten.",
        "",
        "Valores válidos para Tipo: ACTIVO, PASIVO, CAPITAL, INGRESO, GASTO",
        "Valores válidos para Naturaleza: DEUDORA, ACREEDORA",
        "Valores válidos para Nivel: PADRE, DETALLE",
    ]
    for i, txt in enumerate(instrucciones, 1):
        cell = wi.cell(row=i, column=1, value=txt)
        if i == 1:
            cell.font = Font(bold=True, size=14)

    wi.column_dimensions['A'].width = 70

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="plantilla_plan_cuentas.xlsx"'
    wb.save(response)
    return response


@login_required
def cuentas_importar(request):
    """Importar plan de cuentas desde Excel."""
    if request.method != 'POST':
        return redirect('cont-cuentas-lista')

    archivo = request.FILES.get('archivo_excel')
    if not archivo:
        messages.error(request, 'Seleccione un archivo Excel.')
        return redirect('cont-cuentas-lista')

    import openpyxl
    try:
        wb = openpyxl.load_workbook(archivo, read_only=True)
        ws = wb.active

        tipos_validos = {v for v, _ in CuentaContable.TIPO_CHOICES}
        nat_validos = {v for v, _ in CuentaContable.NATURALEZA_CHOICES}
        nivel_validos = {v for v, _ in CuentaContable.NIVEL_CHOICES}

        filas = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        # Primer paso: crear todas las cuentas sin padre
        creadas = 0
        omitidas = 0
        errores = []
        codigos_existentes = set(CuentaContable.objects.values_list('codigo', flat=True))

        # Ordenar por longitud de código (padres primero)
        filas_ordenadas = sorted(filas, key=lambda r: len(str(r[0] or '')))

        for fila in filas_ordenadas:
            if not fila or not fila[0]:
                continue

            codigo = str(fila[0]).strip()
            nombre = str(fila[1] or '').strip()
            tipo = str(fila[2] or '').strip().upper()
            naturaleza = str(fila[3] or '').strip().upper()
            nivel = str(fila[4] or '').strip().upper()
            codigo_padre = str(fila[5] or '').strip() if len(fila) > 5 else ''

            if not codigo or not nombre:
                continue

            if codigo in codigos_existentes:
                omitidas += 1
                continue

            if tipo not in tipos_validos:
                errores.append(f'Fila {codigo}: tipo "{tipo}" inválido')
                continue
            if naturaleza not in nat_validos:
                errores.append(f'Fila {codigo}: naturaleza "{naturaleza}" inválida')
                continue
            if nivel not in nivel_validos:
                errores.append(f'Fila {codigo}: nivel "{nivel}" inválido')
                continue

            padre = None
            if codigo_padre:
                padre = CuentaContable.objects.filter(codigo=codigo_padre).first()

            CuentaContable.objects.create(
                codigo=codigo,
                nombre=nombre,
                tipo=tipo,
                naturaleza=naturaleza,
                nivel=nivel,
                cuenta_padre=padre,
            )
            codigos_existentes.add(codigo)
            creadas += 1

        msg = f'{creadas} cuentas importadas'
        if omitidas:
            msg += f', {omitidas} omitidas (ya existían)'
        if errores:
            msg += f', {len(errores)} errores'
            for e in errores[:5]:
                messages.warning(request, e)
        messages.success(request, msg)

    except Exception as e:
        messages.error(request, f'Error al procesar el archivo: {e}')

    return redirect('cont-cuentas-lista')


@login_required
def cuentas_eliminar(request, pk):
    cuenta = get_object_or_404(CuentaContable, pk=pk)
    if request.method == 'POST':
        if cuenta.lineas.exists():
            messages.error(request, 'No se puede eliminar: la cuenta tiene movimientos registrados.')
        elif cuenta.subcuentas.exists():
            messages.error(request, 'No se puede eliminar: la cuenta tiene subcuentas.')
        else:
            nombre = str(cuenta)
            cuenta.delete()
            messages.success(request, f'Cuenta {nombre} eliminada.')
    return redirect('cont-cuentas-lista')


def _cuenta_desde_post(post, cuenta=None):
    if cuenta is None:
        cuenta = CuentaContable()
    cuenta.codigo      = post.get('codigo', '').strip()
    cuenta.nombre      = post.get('nombre', '').strip()
    cuenta.tipo        = post.get('tipo', '')
    cuenta.naturaleza  = post.get('naturaleza', '')
    cuenta.nivel       = post.get('nivel', 'DETALLE')
    cuenta.descripcion = post.get('descripcion', '').strip()
    cuenta.activa      = post.get('activa') == 'on'
    padre_id           = post.get('cuenta_padre') or None
    cuenta.cuenta_padre_id = padre_id

    if not cuenta.codigo:
        raise ValueError('El código es obligatorio.')
    if not cuenta.nombre:
        raise ValueError('El nombre es obligatorio.')
    return cuenta


# ─────────────────────────────────────────────────────────────────────────────
# ASIENTOS CONTABLES
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def asientos_lista(request):
    q       = request.GET.get('q', '').strip()
    periodo = request.GET.get('periodo', '').strip()
    estado  = request.GET.get('estado', '').strip()

    qs = AsientoContable.objects.all()
    if q:
        qs = qs.filter(Q(concepto__icontains=q) | Q(numero__icontains=q))
    if periodo:
        qs = qs.filter(periodo=periodo)
    if estado:
        qs = qs.filter(estado=estado)

    paginator = Paginator(qs, 20)
    page      = request.GET.get('page')
    asientos  = paginator.get_page(page)

    return render(request, 'contabilidad/asientos/lista.html', {
        'asientos':  asientos,
        'q':         q,
        'f_periodo': periodo,
        'f_estado':  estado,
        'estados':   AsientoContable.ESTADO_CHOICES,
    })


@login_required
def asientos_crear(request):
    cuentas_detalle = CuentaContable.objects.filter(nivel='DETALLE', activa=True).order_by('codigo')

    if request.method == 'POST':
        error = _guardar_asiento(request, asiento=None)
        if error is None:
            messages.success(request, 'Asiento creado correctamente.')
            return redirect('cont-asientos-lista')
        messages.error(request, error)

    return render(request, 'contabilidad/asientos/form.html', {
        'titulo':          'Nuevo asiento',
        'cuentas_detalle': cuentas_detalle,
        'estados':         AsientoContable.ESTADO_CHOICES,
    })


@login_required
def asientos_ver(request, pk):
    asiento = get_object_or_404(AsientoContable, pk=pk)
    lineas  = asiento.lineas.select_related('cuenta').all()
    return render(request, 'contabilidad/asientos/detalle.html', {
        'asiento': asiento,
        'lineas':  lineas,
    })


@login_required
def asientos_editar(request, pk):
    asiento         = get_object_or_404(AsientoContable, pk=pk)
    cuentas_detalle = CuentaContable.objects.filter(nivel='DETALLE', activa=True).order_by('codigo')

    if asiento.estado == 'CONFIRMADO':
        messages.error(request, 'No se puede editar un asiento confirmado.')
        return redirect('cont-asientos-ver', pk=pk)

    if request.method == 'POST':
        error = _guardar_asiento(request, asiento=asiento)
        if error is None:
            messages.success(request, 'Asiento actualizado.')
            return redirect('cont-asientos-ver', pk=pk)
        messages.error(request, error)

    lineas_existentes = list(asiento.lineas.select_related('cuenta').all())
    return render(request, 'contabilidad/asientos/form.html', {
        'titulo':            'Editar asiento',
        'asiento':           asiento,
        'lineas_existentes': lineas_existentes,
        'cuentas_detalle':   cuentas_detalle,
        'estados':           AsientoContable.ESTADO_CHOICES,
    })


@login_required
def asientos_confirmar(request, pk):
    asiento = get_object_or_404(AsientoContable, pk=pk)
    if request.method == 'POST':
        if not asiento.lineas.exists():
            messages.error(request, 'El asiento no tiene líneas.')
        elif not asiento.esta_cuadrado:
            messages.error(request, f'El asiento no cuadra: Debe={asiento.total_debe} / Haber={asiento.total_haber}.')
        else:
            asiento.estado = 'CONFIRMADO'
            asiento.save()
            messages.success(request, f'Asiento #{asiento.numero} confirmado.')
    return redirect('cont-asientos-ver', pk=pk)


@login_required
def asientos_eliminar(request, pk):
    asiento = get_object_or_404(AsientoContable, pk=pk)
    if request.method == 'POST':
        if asiento.estado == 'CONFIRMADO':
            messages.error(request, 'No se puede eliminar un asiento confirmado.')
            return redirect('cont-asientos-ver', pk=pk)
        num = asiento.numero
        asiento.delete()
        messages.success(request, f'Asiento #{num} eliminado.')
    return redirect('cont-asientos-lista')


def _guardar_asiento(request, asiento=None):
    """Guarda un asiento y sus líneas. Retorna None si ok, string de error si falla."""
    post     = request.POST
    fecha    = post.get('fecha', '').strip()
    concepto = post.get('concepto', '').strip()
    estado   = post.get('estado', 'BORRADOR')

    if not fecha or not concepto:
        return 'Fecha y concepto son obligatorios.'

    cuentas_ids = post.getlist('linea_cuenta')
    descrips    = post.getlist('linea_descripcion')
    debes       = post.getlist('linea_debe')
    haberes     = post.getlist('linea_haber')

    lineas_data = []
    for i, cid in enumerate(cuentas_ids):
        if not cid:
            continue
        try:
            cuenta = CuentaContable.objects.get(pk=cid)
        except CuentaContable.DoesNotExist:
            return f'Cuenta con id {cid} no existe.'
        try:
            debe  = Decimal(debes[i]  or '0')
            haber = Decimal(haberes[i] or '0')
        except Exception:
            return 'Valores de debe/haber inválidos.'
        lineas_data.append({
            'cuenta':      cuenta,
            'descripcion': descrips[i] if i < len(descrips) else '',
            'debe':        debe,
            'haber':       haber,
        })

    if not lineas_data:
        return 'El asiento debe tener al menos una línea.'

    total_debe  = sum(l['debe']  for l in lineas_data)
    total_haber = sum(l['haber'] for l in lineas_data)

    if estado == 'CONFIRMADO' and total_debe != total_haber:
        return f'El asiento no cuadra: Debe={total_debe} / Haber={total_haber}.'

    if asiento is None:
        asiento = AsientoContable(creado_por=request.user)
    asiento.fecha    = fecha
    asiento.concepto = concepto
    asiento.estado   = estado
    asiento.save()

    asiento.lineas.all().delete()
    for l in lineas_data:
        LineaAsiento.objects.create(
            asiento=asiento,
            cuenta=l['cuenta'],
            descripcion=l['descripcion'],
            debe=l['debe'],
            haber=l['haber'],
        )
    return None


# ─────────────────────────────────────────────────────────────────────────────
# CUENTAS POR COBRAR
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def cpc_lista(request):
    q       = request.GET.get('q', '').strip()
    estado  = request.GET.get('estado', '').strip()
    vencidas = request.GET.get('vencidas', '').strip()

    qs = CuentaPorCobrar.objects.select_related('receptor', 'factura').all()
    if q:
        qs = qs.filter(Q(receptor__nombre__icontains=q) | Q(factura__numero_control__icontains=q))
    if estado:
        qs = qs.filter(estado=estado)
    if vencidas == '1':
        from datetime import date
        qs = qs.filter(fecha_vencimiento__lt=date.today(), estado__in=['PENDIENTE', 'PARCIAL'])

    paginator = Paginator(qs, 20)
    cuentas   = paginator.get_page(request.GET.get('page'))

    return render(request, 'contabilidad/cpc/lista.html', {
        'cuentas':  cuentas,
        'q':        q,
        'f_estado': estado,
        'f_vencidas': vencidas,
        'estados':  CuentaPorCobrar.ESTADO_CHOICES,
    })


@login_required
def cpc_crear(request):
    """Selecciona una FacturaElectronica y registra como Cuenta por Cobrar."""
    # Facturas que aún no tienen CxC
    facturas = FacturaElectronica.objects.exclude(
        cuentas_cobrar__isnull=False
    ).select_related('dtereceptor').order_by('-fecha_emision')[:200]

    if request.method == 'POST':
        factura_id       = request.POST.get('factura')
        fecha_vencimiento = request.POST.get('fecha_vencimiento')
        notas            = request.POST.get('notas', '').strip()

        try:
            factura = FacturaElectronica.objects.get(pk=factura_id)
        except FacturaElectronica.DoesNotExist:
            messages.error(request, 'Factura no encontrada.')
            return redirect('cont-cpc-crear')

        if CuentaPorCobrar.objects.filter(factura=factura).exists():
            messages.error(request, 'Esta factura ya tiene una cuenta por cobrar registrada.')
            return redirect('cont-cpc-lista')

        cpc = CuentaPorCobrar.objects.create(
            factura=factura,
            receptor=factura.dtereceptor,
            fecha_emision=factura.fecha_emision,
            fecha_vencimiento=fecha_vencimiento,
            monto_original=factura.total_pagar,
            notas=notas,
            creado_por=request.user,
        )
        messages.success(request, f'Cuenta por Cobrar #{cpc.pk} creada correctamente.')
        return redirect('cont-cpc-detalle', pk=cpc.pk)

    return render(request, 'contabilidad/cpc/crear.html', {'facturas': facturas})


@login_required
def cpc_detalle(request, pk):
    cpc   = get_object_or_404(CuentaPorCobrar, pk=pk)
    pagos = cpc.pagos.select_related('asiento', 'cuenta_debito', 'cuenta_credito').all()
    cuentas_detalle = CuentaContable.objects.filter(nivel='DETALLE', activa=True).order_by('codigo')
    return render(request, 'contabilidad/cpc/detalle.html', {
        'cpc':             cpc,
        'pagos':           pagos,
        'cuentas_detalle': cuentas_detalle,
    })


@login_required
def cpc_registrar_pago(request, pk):
    cpc = get_object_or_404(CuentaPorCobrar, pk=pk)

    if cpc.estado == 'PAGADO':
        messages.error(request, 'Esta cuenta ya está completamente pagada.')
        return redirect('cont-cpc-detalle', pk=pk)
    if cpc.estado == 'ANULADO':
        messages.error(request, 'No se puede registrar pago en una cuenta anulada.')
        return redirect('cont-cpc-detalle', pk=pk)

    if request.method == 'POST':
        try:
            monto = Decimal(request.POST.get('monto', '0'))
        except Exception:
            messages.error(request, 'Monto inválido.')
            return redirect('cont-cpc-detalle', pk=pk)

        if monto <= 0:
            messages.error(request, 'El monto debe ser mayor a cero.')
            return redirect('cont-cpc-detalle', pk=pk)
        if monto > cpc.saldo_pendiente:
            messages.error(request, f'El monto excede el saldo pendiente (${cpc.saldo_pendiente}).')
            return redirect('cont-cpc-detalle', pk=pk)

        cuenta_debito_id  = request.POST.get('cuenta_debito')
        cuenta_credito_id = request.POST.get('cuenta_credito')

        pago = PagoCobro(
            cuenta_cobrar=cpc,
            fecha=request.POST.get('fecha'),
            monto=monto,
            forma_pago=request.POST.get('forma_pago', 'EFECTIVO'),
            referencia=request.POST.get('referencia', '').strip(),
            notas=request.POST.get('notas', '').strip(),
            creado_por=request.user,
        )
        if cuenta_debito_id:
            pago.cuenta_debito_id = cuenta_debito_id
        if cuenta_credito_id:
            pago.cuenta_credito_id = cuenta_credito_id
        pago.save()

        if pago.cuenta_debito and pago.cuenta_credito:
            pago.generar_asiento()
            messages.success(request, f'Pago de ${monto} registrado con asiento contable generado.')
        else:
            messages.success(request, f'Pago de ${monto} registrado. Sin asiento (no se seleccionaron cuentas).')

        cpc.actualizar_estado()

    return redirect('cont-cpc-detalle', pk=pk)


@login_required
def cpc_anular(request, pk):
    cpc = get_object_or_404(CuentaPorCobrar, pk=pk)
    if request.method == 'POST':
        if cpc.estado == 'PAGADO':
            messages.error(request, 'No se puede anular una cuenta completamente pagada.')
        else:
            cpc.estado = 'ANULADO'
            cpc.save(update_fields=['estado'])
            messages.success(request, f'Cuenta por Cobrar #{cpc.pk} anulada.')
    return redirect('cont-cpc-lista')


# ─────────────────────────────────────────────────────────────────────────────
# CUENTAS POR PAGAR
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def cpp_lista(request):
    q       = request.GET.get('q', '').strip()
    estado  = request.GET.get('estado', '').strip()
    vencidas = request.GET.get('vencidas', '').strip()

    qs = CuentaPorPagar.objects.select_related('proveedor', 'compra').all()
    if q:
        qs = qs.filter(Q(proveedor__nombre__icontains=q) | Q(compra__numero_documento__icontains=q))
    if estado:
        qs = qs.filter(estado=estado)
    if vencidas == '1':
        from datetime import date
        qs = qs.filter(fecha_vencimiento__lt=date.today(), estado__in=['PENDIENTE', 'PARCIAL'])

    paginator = Paginator(qs, 20)
    cuentas   = paginator.get_page(request.GET.get('page'))

    return render(request, 'contabilidad/cpp/lista.html', {
        'cuentas':    cuentas,
        'q':          q,
        'f_estado':   estado,
        'f_vencidas': vencidas,
        'estados':    CuentaPorPagar.ESTADO_CHOICES,
    })


@login_required
def cpp_crear(request):
    """Selecciona una Compra y registra como Cuenta por Pagar."""
    compras = Compra.objects.exclude(
        cuentas_pagar__isnull=False
    ).select_related('proveedor').order_by('-fecha')[:200]

    if request.method == 'POST':
        compra_id        = request.POST.get('compra')
        fecha_vencimiento = request.POST.get('fecha_vencimiento')
        notas            = request.POST.get('notas', '').strip()

        try:
            compra = Compra.objects.get(pk=compra_id)
        except Compra.DoesNotExist:
            messages.error(request, 'Compra no encontrada.')
            return redirect('cont-cpp-crear')

        if CuentaPorPagar.objects.filter(compra=compra).exists():
            messages.error(request, 'Esta compra ya tiene una cuenta por pagar registrada.')
            return redirect('cont-cpp-lista')

        cpp = CuentaPorPagar.objects.create(
            compra=compra,
            proveedor=compra.proveedor,
            fecha_emision=compra.fecha.date() if hasattr(compra.fecha, 'date') else compra.fecha,
            fecha_vencimiento=fecha_vencimiento,
            monto_original=compra.total,
            notas=notas,
            creado_por=request.user,
        )
        messages.success(request, f'Cuenta por Pagar #{cpp.pk} creada correctamente.')
        return redirect('cont-cpp-detalle', pk=cpp.pk)

    return render(request, 'contabilidad/cpp/crear.html', {'compras': compras})


@login_required
def cpp_detalle(request, pk):
    cpp   = get_object_or_404(CuentaPorPagar, pk=pk)
    pagos = cpp.pagos.select_related('asiento', 'cuenta_debito', 'cuenta_credito').all()
    cuentas_detalle = CuentaContable.objects.filter(nivel='DETALLE', activa=True).order_by('codigo')
    return render(request, 'contabilidad/cpp/detalle.html', {
        'cpp':             cpp,
        'pagos':           pagos,
        'cuentas_detalle': cuentas_detalle,
    })


@login_required
def cpp_registrar_pago(request, pk):
    cpp = get_object_or_404(CuentaPorPagar, pk=pk)

    if cpp.estado == 'PAGADO':
        messages.error(request, 'Esta cuenta ya está completamente pagada.')
        return redirect('cont-cpp-detalle', pk=pk)
    if cpp.estado == 'ANULADO':
        messages.error(request, 'No se puede registrar pago en una cuenta anulada.')
        return redirect('cont-cpp-detalle', pk=pk)

    if request.method == 'POST':
        try:
            monto = Decimal(request.POST.get('monto', '0'))
        except Exception:
            messages.error(request, 'Monto inválido.')
            return redirect('cont-cpp-detalle', pk=pk)

        if monto <= 0:
            messages.error(request, 'El monto debe ser mayor a cero.')
            return redirect('cont-cpp-detalle', pk=pk)
        if monto > cpp.saldo_pendiente:
            messages.error(request, f'El monto excede el saldo pendiente (${cpp.saldo_pendiente}).')
            return redirect('cont-cpp-detalle', pk=pk)

        cuenta_debito_id  = request.POST.get('cuenta_debito')
        cuenta_credito_id = request.POST.get('cuenta_credito')

        pago = PagoPagar(
            cuenta_pagar=cpp,
            fecha=request.POST.get('fecha'),
            monto=monto,
            forma_pago=request.POST.get('forma_pago', 'EFECTIVO'),
            referencia=request.POST.get('referencia', '').strip(),
            notas=request.POST.get('notas', '').strip(),
            creado_por=request.user,
        )
        if cuenta_debito_id:
            pago.cuenta_debito_id = cuenta_debito_id
        if cuenta_credito_id:
            pago.cuenta_credito_id = cuenta_credito_id
        pago.save()

        if pago.cuenta_debito and pago.cuenta_credito:
            pago.generar_asiento()
            messages.success(request, f'Pago de ${monto} registrado con asiento contable generado.')
        else:
            messages.success(request, f'Pago de ${monto} registrado. Sin asiento (no se seleccionaron cuentas).')

        cpp.actualizar_estado()

    return redirect('cont-cpp-detalle', pk=pk)


@login_required
def cpp_anular(request, pk):
    cpp = get_object_or_404(CuentaPorPagar, pk=pk)
    if request.method == 'POST':
        if cpp.estado == 'PAGADO':
            messages.error(request, 'No se puede anular una cuenta completamente pagada.')
        else:
            cpp.estado = 'ANULADO'
            cpp.save(update_fields=['estado'])
            messages.success(request, f'Cuenta por Pagar #{cpp.pk} anulada.')
    return redirect('cont-cpp-lista')


# ─────────────────────────────────────────────────────────────────────────────
# REPORTES CONTABLES
# ─────────────────────────────────────────────────────────────────────────────

def _filtrar_periodo(request):
    """Extrae fecha_desde / fecha_hasta del GET y devuelve (desde, hasta)."""
    desde = request.GET.get('desde', '')
    hasta = request.GET.get('hasta', '')
    return desde, hasta


def _saldos_cuentas(desde, hasta):
    """Devuelve queryset de CuentaContable con debe_total, haber_total y saldo."""
    filtro = Q(lineas__asiento__estado='CONFIRMADO')
    if desde:
        filtro &= Q(lineas__asiento__fecha__gte=desde)
    if hasta:
        filtro &= Q(lineas__asiento__fecha__lte=hasta)

    cuentas = CuentaContable.objects.filter(filtro).annotate(
        debe_total=Sum('lineas__debe', default=0),
        haber_total=Sum('lineas__haber', default=0),
    ).order_by('codigo')

    for c in cuentas:
        if c.naturaleza == 'DEUDORA':
            c.saldo_calc = c.debe_total - c.haber_total
        else:
            c.saldo_calc = c.haber_total - c.debe_total
    return cuentas


@login_required
def libro_mayor(request):
    desde, hasta = _filtrar_periodo(request)
    cuenta_id = request.GET.get('cuenta', '')

    cuentas_all = CuentaContable.objects.filter(nivel='DETALLE').order_by('codigo')
    lineas = LineaAsiento.objects.filter(
        asiento__estado='CONFIRMADO'
    ).select_related('cuenta', 'asiento').order_by('asiento__fecha', 'asiento__numero')

    if cuenta_id:
        lineas = lineas.filter(cuenta_id=cuenta_id)
    if desde:
        lineas = lineas.filter(asiento__fecha__gte=desde)
    if hasta:
        lineas = lineas.filter(asiento__fecha__lte=hasta)

    # Agrupar por cuenta
    cuentas_dict = {}
    for l in lineas:
        cid = l.cuenta_id
        if cid not in cuentas_dict:
            cuentas_dict[cid] = {
                'cuenta': l.cuenta,
                'lineas': [],
                'total_debe': Decimal(0),
                'total_haber': Decimal(0),
            }
        cuentas_dict[cid]['lineas'].append(l)
        cuentas_dict[cid]['total_debe'] += l.debe
        cuentas_dict[cid]['total_haber'] += l.haber

    for v in cuentas_dict.values():
        c = v['cuenta']
        if c.naturaleza == 'DEUDORA':
            v['saldo'] = v['total_debe'] - v['total_haber']
        else:
            v['saldo'] = v['total_haber'] - v['total_debe']

    return render(request, 'contabilidad/libro_mayor.html', {
        'cuentas_data': sorted(cuentas_dict.values(), key=lambda x: x['cuenta'].codigo),
        'cuentas_all': cuentas_all,
        'cuenta_sel': cuenta_id,
        'desde': desde,
        'hasta': hasta,
    })


@login_required
def balance_comprobacion(request):
    desde, hasta = _filtrar_periodo(request)
    cuentas = _saldos_cuentas(desde, hasta)

    total_debe = sum(c.debe_total for c in cuentas)
    total_haber = sum(c.haber_total for c in cuentas)
    total_saldo_deudor = sum(c.saldo_calc for c in cuentas if c.saldo_calc > 0)
    total_saldo_acreedor = sum(abs(c.saldo_calc) for c in cuentas if c.saldo_calc < 0)

    return render(request, 'contabilidad/balance_comprobacion.html', {
        'cuentas': cuentas,
        'total_debe': total_debe,
        'total_haber': total_haber,
        'total_saldo_deudor': total_saldo_deudor,
        'total_saldo_acreedor': total_saldo_acreedor,
        'desde': desde,
        'hasta': hasta,
    })


@login_required
def balance_general(request):
    desde, hasta = _filtrar_periodo(request)
    cuentas = _saldos_cuentas(desde, hasta)

    activos = [c for c in cuentas if c.tipo == 'ACTIVO']
    pasivos = [c for c in cuentas if c.tipo == 'PASIVO']
    capital = [c for c in cuentas if c.tipo == 'CAPITAL']

    total_activos = sum(c.saldo_calc for c in activos)
    total_pasivos = sum(c.saldo_calc for c in pasivos)
    total_capital = sum(c.saldo_calc for c in capital)

    # Resultado del periodo (ingresos - gastos)
    ingresos = [c for c in cuentas if c.tipo == 'INGRESO']
    gastos = [c for c in cuentas if c.tipo == 'GASTO']
    resultado = sum(c.saldo_calc for c in ingresos) - sum(c.saldo_calc for c in gastos)

    return render(request, 'contabilidad/balance_general.html', {
        'activos': activos,
        'pasivos': pasivos,
        'capital': capital,
        'total_activos': total_activos,
        'total_pasivos': total_pasivos,
        'total_capital': total_capital,
        'resultado': resultado,
        'desde': desde,
        'hasta': hasta,
    })


@login_required
def estado_resultados(request):
    desde, hasta = _filtrar_periodo(request)
    cuentas = _saldos_cuentas(desde, hasta)

    ingresos = [c for c in cuentas if c.tipo == 'INGRESO']
    gastos = [c for c in cuentas if c.tipo == 'GASTO']

    total_ingresos = sum(c.saldo_calc for c in ingresos)
    total_gastos = sum(c.saldo_calc for c in gastos)
    utilidad = total_ingresos - total_gastos

    return render(request, 'contabilidad/estado_resultados.html', {
        'ingresos': ingresos,
        'gastos': gastos,
        'total_ingresos': total_ingresos,
        'total_gastos': total_gastos,
        'utilidad': utilidad,
        'desde': desde,
        'hasta': hasta,
    })


@login_required
def reportes_anexos(request):
    return render(request, 'contabilidad/reportes_anexos.html')