# =========================
#   VENTAS (front-office)
# =========================
from datetime import datetime, date, timedelta
from django.contrib import messages
from django.db import transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponseBadRequest
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from django.db.models import Q, F, Sum, Count
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from django.views.decorators.http import require_POST
import requests
from django.utils import timezone
from django.views.decorators.http import require_http_methods
#from django.contrib.auth.models import User  # si usas auth default
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.template.loader import render_to_string
try:
    from weasyprint import HTML
    WEASYPRINT_OK = True
except Exception as e:
    HTML = None
    WEASYPRINT_OK = False

from django.utils.encoding import smart_str

from django.db.models import Sum, Case, When, Prefetch, FloatField, F, ExpressionWrapper, Value, DecimalField, Q
from math import radians, cos, sin, asin, sqrt

from django.db.models.functions import Coalesce

from django.utils.timezone import now

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Modelos FE
from .models import ActividadEconomica, NumeroControl, Pais, Receptor_fe, TipoPersona, TiposDocIDReceptor, Municipio, Tipo_dte, Modelofacturacion, TipoTransmision
from .models import FacturaElectronica, DetalleFactura, Descuento
# Modelos INVENTARIO
from INVENTARIO.models import Producto, MovimientoInventario, DevolucionVenta, DetalleDevolucionVenta, Almacen, Tributo, TipoItem, TipoUnidadMedida

CART_SESSION_KEY = "cart_by_receptor"    # { "<receptor_id>": { "<producto_id>": {qty, precio, iva_on, desc_pct} } }

from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from django.db import transaction
from django.db.models import Q
import re
from django.contrib.auth import get_user_model

def _none_if_blank(v):
    v = (v or "").strip()
    return v or None

User = get_user_model()

def _get_cart(request):
    cart = request.session.get(CART_SESSION_KEY) or {}
    if not isinstance(cart, dict):
        cart = {}
    return cart

def _save_cart(request, cart):
    request.session[CART_SESSION_KEY] = cart
    request.session.modified = True

def _ensure_receptor(receptor_id):
    if not receptor_id:
        return None
    try:
        return Receptor_fe.objects.get(id=receptor_id)
    except Receptor_fe.DoesNotExist:
        return None
# ------------- EXPORTACIONES -------------

@login_required
def exportar_facturas_excel(request):
    # aplicar mismos filtros que en tu lista
    qs = FacturaElectronica.objects.select_related("dtereceptor", "tipo_dte").all()

    if request.GET.get("recibido_mh"):
        qs = qs.filter(recibido_mh=(request.GET["recibido_mh"] == "True"))
    if request.GET.get("tipo_dte"):
        qs = qs.filter(tipo_dte_id=request.GET["tipo_dte"])
    if request.GET.get("sello_recepcion"):
        qs = qs.filter(sello_recepcion__icontains=request.GET["sello_recepcion"])

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas"

    # Encabezados
    headers = ["Número de Control", "Cliente", "Tipo DTE", "Fecha", "Total", "IVA", "Recibido MH", "Estado"]
    ws.append(headers)

    for factura in qs:
        ws.append([
            factura.numero_control,
            factura.dtereceptor.nombre if factura.dtereceptor else "",
            factura.tipo_dte.descripcion if factura.tipo_dte else "",
            factura.fecha_emision.strftime("%Y-%m-%d") if factura.fecha_emision else "",
            float(factura.total_pagar or 0),
            float(factura.total_iva or 0),
            "Sí" if factura.recibido_mh else "No",
            "Anulada" if factura.dte_invalidacion.exists() else "Viva"
        ])

    # Ajustar ancho de columnas
    for col_num, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    # Preparar respuesta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="facturas.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_facturas_pdf(request):
    qs = FacturaElectronica.objects.select_related(
        "dtereceptor", "tipo_dte"
    ).prefetch_related("detalles__producto")

    # filtros
    if request.GET.get("recibido_mh"):
        qs = qs.filter(recibido_mh=(request.GET["recibido_mh"] == "True"))
    if request.GET.get("tipo_dte"):
        qs = qs.filter(tipo_dte_id=request.GET["tipo_dte"])

    fecha_ini = request.GET.get("fecha_ini") or None
    fecha_fin = request.GET.get("fecha_fin") or None

    if fecha_ini:
        qs = qs.filter(fecha_emision__gte=fecha_ini)
    if fecha_fin:
        qs = qs.filter(fecha_emision__lte=fecha_fin)

    # totales
    total_facturas = qs.count()
    total_monto = qs.aggregate(s=Sum("total_pagar"))["s"] or 0
    total_iva = qs.aggregate(s=Sum("total_iva"))["s"] or 0

    from .models import DetalleFactura
    detalles = DetalleFactura.objects.filter(factura__in=qs)
    total_lineas = detalles.count()
    total_cantidad = detalles.aggregate(s=Sum("cantidad"))["s"] or 0

    # 🆕 enviar las fechas al template
    html_string = render_to_string("reportes/facturas_pdf.html", {
        "facturas": qs,
        "total_facturas": total_facturas,
        "total_monto": total_monto,
        "total_iva": total_iva,
        "total_lineas": total_lineas,
        "total_cantidad": total_cantidad,
        "fecha_ini": fecha_ini,
        "fecha_fin": fecha_fin,
    })
    pdf_file = HTML(string=html_string).write_pdf()

    response = HttpResponse(pdf_file, content_type="application/pdf")
    response["Content-Disposition"] = 'inline; filename="facturas.pdf"'
    return response

# ---------- HOME VENTAS ----------
def _pick_field(model, candidates):
    """Devuelve el primer field existente en el modelo."""
    if not model:
        return None
    names = {f.name for f in model._meta.get_fields()}
    for name in candidates:
        if name in names:
            return name
    return None


@login_required
def ventas_home(request):
    today = timezone.localdate()
    fac_date_field = _pick_field(
        FacturaElectronica,
        ["fecha_emision", "fecha", "fecha_generacion", "created_at", "created", "fecha_creacion"]
    )

    ventas_hoy_total = Decimal("0.00")
    ventas_hoy_count = 0

    if fac_date_field:
        filtro = {fac_date_field: today}
        qs = FacturaElectronica.objects.filter(**filtro)
        ventas_hoy_total = qs.aggregate(
            s=Coalesce(Sum("total_pagar"), Decimal("0.00"))
        )["s"] or Decimal("0.00")
        ventas_hoy_count = qs.count()

    context = {
        "ventas_hoy_total": ventas_hoy_total,
        "ventas_hoy_count": ventas_hoy_count,
        "today": now().date().isoformat(),
    }
    return render(request, "ventas/home.html", context)

# --------- CLIENTES (Receptor_fe) ----------

# ================= Helpers =================


DUI_REGEX = re.compile(r"^[0-9]{8}-[0-9]$")

def _ensure_id_list(val):
    """
    Devuelve una lista de strings con IDs a partir de:
    - None / ''      -> []
    - int / str      -> ['<val>']
    - list/tuple/set -> [str(...), ...]
    """
    if val is None or val == '':
        return []
    if isinstance(val, (list, tuple, set)):
        return [str(v) for v in val if v not in (None, '')]
    return [str(val)]

def _none_if_blank(v):
    v = (v or "").strip()
    return v or None

def _parse_coord(v):
    try:
        if v is None or str(v).strip() == "":
            return None
        return Decimal(str(v))
    except Exception:
        return None

def _geocode_address_osm(query: str):
    """
    Stub de geocodificación. Devuelve (lat, lng, source)
    Implementa aquí tu consulta real si la tienes.
    """
    return (None, None, "")

# Defaults “quemados” para NATURAL
DEFAULTS_NAT = {
    "correo": "ventasdetalleincoe@gmail.com",
    "direccion": "N/A",
    "pais_id": 11,      # <-- pon el ID real de EL SALVADOR en tu tabla País
    "actividad_id": 771,  # <-- pon el ID de tu actividad default
}

def _get_dui_tipo_doc():
    # Intenta por codigo="2" o por texto "DUI"
    td = TiposDocIDReceptor.objects.filter(codigo="2").first()
    if td:
        return td
    return TiposDocIDReceptor.objects.filter(descripcion__icontains="DUI").first()

def _parse_coord(val):
    try:
        if val in (None, "", "null"):
            return None
        # Normalizar: convertir comas decimales a puntos
        val = str(val).strip().replace(",", ".")
        # Convertir a Decimal manteniendo todos los decimales
        return Decimal(val)
    except (InvalidOperation, ValueError, TypeError):
        return None
     
@login_required
def _geocode_address_osm(q):
    """
    Geocodifica con Nominatim (OSM).
    Respeta su política: añade User-Agent propio.
    """
    try:
        r = requests.get(
            'https://nominatim.openstreetmap.org/search',
            params={'q': q, 'format': 'json', 'addressdetails': 0, 'limit': 1},
            headers={'User-Agent': 'INTRACIOE/1.0 (+https://intracioe.com)'},
            timeout=5,
            )
        r.raise_for_status()
        data = r.json()
        if data:
            return Decimal(data[0]['lat']), Decimal(data[0]['lon']), "nominatim"
    except Exception:
        pass
    return None, None, ""


def haversine(lat1, lon1, lat2, lon2):
    """
    Calcula la distancia en km entre dos puntos (lat/lng)
    """
    # Aseguramos que todo sea float
    lat1, lon1, lat2, lon2 = map(float, [lat1, lon1, lat2, lon2])

    R = 6378  # Radio de la Tierra en km
    dlat = radians(lat2 - lat1)
    dlon = radians(lon2 - lon1)
    a = sin(dlat/2)**2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon/2)**2
    c = 2 * asin(sqrt(a))
    return R * c

@login_required
def _cliente_form(request, c: Receptor_fe | None = None):
    tipos_doc = TiposDocIDReceptor.objects.all().order_by('descripcion')
    municipios = Municipio.objects.select_related('departamento').order_by('descripcion')
    paises = Pais.objects.all().order_by('descripcion')
    tipos_persona = TipoPersona.objects.all().order_by('descripcion')
    actividades = ActividadEconomica.objects.all().order_by('descripcion')

    if request.method == "POST":
        tipo_persona_id = request.POST.get('tipo_persona')
        if not tipo_persona_id:
            messages.error(request, "Selecciona el tipo de persona.")
            return redirect(request.path if not c else request.path)

        tipo_persona = get_object_or_404(TipoPersona, pk=tipo_persona_id)

        # comunes
        nombre      = _none_if_blank(request.POST.get('nombre'))
        telefono    = _none_if_blank(request.POST.get('telefono'))
        correo      = _none_if_blank(request.POST.get('correo'))
        nombre_com  = _none_if_blank(request.POST.get('nombreComercial'))
        direccion   = _none_if_blank(request.POST.get('direccion'))
        municipio_id= request.POST.get('municipio')
        pais_id     = request.POST.get('pais')
        lat         = _parse_coord(request.POST.get('lat'))
        lng         = _parse_coord(request.POST.get('lng'))

        # NATURAL
        if str(getattr(tipo_persona, 'codigo', tipo_persona_id)) == "1":
            # DUI por defecto
            tipo_doc = (TiposDocIDReceptor.objects.filter(codigo="13").first()
                        or TiposDocIDReceptor.objects.filter(descripcion__icontains="DUI").first())
            if not tipo_doc:
                messages.error(request, "No está configurado el tipo de documento DUI.")
                return redirect(request.path)

            num_documento = _none_if_blank(request.POST.get('num_documento'))
            if not num_documento or not DUI_REGEX.match(num_documento):
                messages.error(request, "Ingrese un DUI válido con formato ########-#.")
                return redirect(request.path)

            if not (nombre and telefono):
                messages.error(request, "Nombre y teléfono son obligatorios para persona natural.")
                return redirect(request.path)

            # defaults
            if not correo:    correo = DEFAULTS_NAT["correo"]
            if not direccion: direccion = DEFAULTS_NAT["direccion"]
            pais = Pais.objects.filter(pk=DEFAULTS_NAT["pais_id"]).first()
            municipio = Municipio.objects.filter(pk=municipio_id).first() if municipio_id else None

            geocode_source = ""; geocoded_at = None
            if municipio and (lat is None or lng is None) and direccion:
                muni = municipio.descripcion or ""
                depto = getattr(municipio.departamento, "descripcion", "") or ""
                q = f"{direccion}, {muni}, {depto}, El Salvador"
                lat, lng, geocode_source = _geocode_address_osm(q)
                if lat is not None and lng is not None:
                    geocoded_at = timezone.now()
            elif lat is not None and lng is not None:
                geocode_source = "manual"; geocoded_at = timezone.now()

            with transaction.atomic():
                if c is None:
                    c = Receptor_fe.objects.create(
                        tipo_documento=tipo_doc,
                        num_documento=num_documento,
                        nrc=None,
                        nombre=nombre,
                        municipio=municipio,
                        direccion=direccion,
                        telefono=telefono,
                        correo=correo,
                        nombreComercial=nombre_com or nombre,
                        pais=pais,
                        tipo_persona=tipo_persona,
                        lat=lat, lng=lng, geocode_source=geocode_source, geocoded_at=geocoded_at,
                    )
                else:
                    c.tipo_documento = tipo_doc
                    c.num_documento  = num_documento
                    c.nrc=None
                    c.nombre=nombre
                    c.municipio=municipio
                    c.direccion=direccion
                    c.telefono=telefono
                    c.correo=correo
                    c.nombreComercial=nombre_com or nombre
                    c.pais=pais
                    c.tipo_persona=tipo_persona
                    c.lat=lat; c.lng=lng
                    c.geocode_source=geocode_source; c.geocoded_at=geocoded_at
                    c.save()

                # Actividad por defecto
                raw_act = request.POST.get('actividades_economicas') or DEFAULTS_NAT.get("actividad_id")
                act_ids = _ensure_id_list(raw_act)

                if act_ids:
                    acts_qs = ActividadEconomica.objects.filter(pk__in=act_ids)
                    c.actividades_economicas.set(acts_qs)
                else:
                    c.actividades_economicas.clear()

            messages.success(request, f"Cliente {'creado' if request.resolver_match.url_name=='clientes_crear' else 'actualizado'} (Persona Natural).")
            return redirect('clientes_list')

        # JURÍDICA
        else:
            tipo_doc_id   = request.POST.get('tipo_documento')
            num_documento = _none_if_blank(request.POST.get('num_documento'))
            nrc           = _none_if_blank(request.POST.get('nrc'))
            act_ids       = request.POST.getlist('actividades_economicas') or request.POST.getlist('actividades')

            if not (tipo_doc_id and num_documento and nombre and municipio_id):
                messages.error(request, "Completa tipo documento, número, nombre y municipio.")
                return redirect(request.path)

            tipo_doc = get_object_or_404(TiposDocIDReceptor, pk=tipo_doc_id)
            municipio = get_object_or_404(Municipio, pk=municipio_id)
            pais = Pais.objects.filter(pk=pais_id).first() if pais_id else None

            geocode_source=""; geocoded_at=None
            if (lat is None or lng is None) and direccion:
                muni = municipio.descripcion or ""
                depto = getattr(municipio.departamento, "descripcion", "") or ""
                q = f"{direccion}, {muni}, {depto}, El Salvador"
                lat, lng, geocode_source = _geocode_address_osm(q)
                if lat is not None and lng is not None:
                    geocoded_at = timezone.now()
            elif lat is not None and lng is not None:
                geocode_source="manual"; geocoded_at=timezone.now()

            with transaction.atomic():
                if c is None:
                    c = Receptor_fe.objects.create(
                        tipo_documento=tipo_doc,
                        num_documento=num_documento,
                        nrc=nrc,
                        nombre=nombre,
                        municipio=municipio,
                        direccion=direccion,
                        telefono=telefono,
                        correo=correo,
                        nombreComercial=nombre_com or nombre,
                        pais=pais,
                        tipo_persona=tipo_persona,
                        lat=lat, lng=lng, geocode_source=geocode_source, geocoded_at=geocoded_at,
                    )
                else:
                    c.tipo_documento=tipo_doc
                    c.num_documento=num_documento
                    c.nrc=nrc
                    c.nombre=nombre
                    c.municipio=municipio
                    c.direccion=direccion
                    c.telefono=telefono
                    c.correo=correo
                    c.nombreComercial=nombre_com or nombre
                    c.pais=pais
                    c.tipo_persona=tipo_persona
                    c.lat=lat; c.lng=lng
                    c.geocode_source=geocode_source; c.geocoded_at=geocoded_at
                    c.save()

                raw_act_ids = request.POST.getlist('actividades_economicas') or request.POST.getlist('actividades')
                act_ids = _ensure_id_list(raw_act_ids)

                if act_ids:
                    acts_qs = ActividadEconomica.objects.filter(pk__in=act_ids)
                    c.actividades_economicas.set(acts_qs)
                else:
                    c.actividades_economicas.clear()

            messages.success(request, f"Cliente {'creado' if request.resolver_match.url_name=='clientes_crear' else 'actualizado'} (Persona Jurídica).")
            return redirect('clientes_list')

    # GET
    return render(request, 'ventas/clientes/form.html', {
        'c': c,
        'tipos_doc': tipos_doc,
        'municipios': municipios,
        'paises': paises,
        'tipos_persona': tipos_persona,
        'actividades': actividades,
    })


# ===== LISTA =====
@login_required
def clientes_list(request):
    q   = request.GET.get('q', '').strip()
    lat = request.GET.get('lat')
    lng = request.GET.get('lng')

    clientes = Receptor_fe.objects.select_related('municipio','municipio__departamento').all()

    if q:
        clientes = clientes.filter(
            Q(nombre__icontains=q) |
            Q(num_documento__icontains=q) |
            Q(telefono__icontains=q) |
            Q(correo__icontains=q)
        ).order_by('nombre')
    else:
        clientes = clientes.order_by('nombre')

    # geo opcional
    lat_user = lng_user = None
    if lat and lng:
        try:
            from math import radians, sin, cos, asin, sqrt
            def haversine(lat1, lon1, lat2, lon2):
                R = 6371.0
                dlat = radians(lat2-lat1); dlon = radians(lon2-lon1)
                a = sin(dlat/2)**2 + cos(radians(lat1))*cos(radians(lat2))*sin(dlon/2)**2
                return 2*R*asin(sqrt(a))

            lat_user = float(lat); lng_user = float(lng)
            tmp = []
            for c in clientes:
                if c.lat and c.lng:
                    c.distancia_km = round(haversine(float(lat_user), float(lng_user), float(c.lat), float(c.lng)), 2)
                else:
                    c.distancia_km = None
                tmp.append(c)
            clientes = sorted(tmp, key=lambda x: x.distancia_km if x.distancia_km is not None else 9e9)
        except ValueError:
            pass

    paginator = Paginator(clientes, 10)
    page_number = request.GET.get('page')
    clientes_page = paginator.get_page(page_number)

    return render(request, 'ventas/clientes/list.html', {
        'clientes': clientes_page,
        'q': q,
        'lat': lat_user,
        'lng': lng_user,
        'carrito_count': 0,
    })


@login_required
@require_http_methods(["GET","POST"])
def clientes_crear(request):
    return _cliente_form(request)

@login_required
@require_POST
def clientes_eliminar(request, pk):
    Receptor_fe.objects.filter(pk=pk).delete()
    messages.success(request, "Cliente eliminado.")
    return redirect('clientes_list')


@login_required
@require_http_methods(["GET","POST"])
def clientes_editar(request, pk):
    c = get_object_or_404(Receptor_fe, pk=pk)
    return _cliente_form(request, c)

# ---------- CATÁLOGO ----------
@login_required
def catalogo_productos(request):
    q = (request.GET.get('q') or '').strip()
    receptor_id = request.GET.get('receptor_id') or ''
    productos = Producto.objects.all()
    if q:
        productos = productos.filter(
            Q(codigo__icontains=q) | Q(descripcion__icontains=q) | Q(nombre__icontains=q)
        )

    context = {
        "q": q,
        "productos": productos,
        "receptores": Receptor_fe.objects.values("id","num_documento","nombre"),
        "receptor_actual": receptor_id,
    }
    return render(request, "ventas/catalogo.html", context)

# ---------- CARRITO por cliente (session) ----------
def _cart_key(receptor_id: int) -> str:
    return f'cart:{receptor_id}'

@login_required
def carrito_ver(request):
    """Pantalla del carrito (elige cliente, agrega productos, edita cantidades)."""
    rid = request.GET.get("receptor_id") or ""
    receptor = _ensure_receptor(rid)
    cart = _get_cart(request)
    items = []
    total = Decimal("0.00")
    if receptor:
        bucket = cart.get(str(receptor.id), {})
        # cargar datos “vivos” del producto (nombre / precio si no lo guardaste)
        for pid, row in bucket.items():
            try:
                p = Producto.objects.get(id=pid)
            except Producto.DoesNotExist:
                continue
            qty = int(row.get("qty", 1))
            precio = Decimal(str(row.get("precio"))) if row.get("precio") not in (None, "") else Decimal(str(p.preunitario or "0"))
            iva_on = bool(row.get("iva_on", False))
            desc = Decimal(str(row.get("desc_pct", "0")))
            base = (precio * qty)
            mto_desc = (base * (desc/Decimal("100"))).quantize(Decimal("0.01"))
            base_neta = base - mto_desc
            iva = (base_neta * Decimal("0.13")).quantize(Decimal("0.01")) if iva_on else Decimal("0.00")
            linea = (base_neta + iva).quantize(Decimal("0.01"))

            items.append({
                "id": p.id, "codigo": p.codigo, "nombre": p.descripcion,
                "qty": qty, "precio": f"{precio:.2f}", "iva_on": iva_on, "desc_pct": f"{desc:.2f}",
                "total": f"{linea:.2f}"
            })
            total += linea

    context = {
        "receptores": Receptor_fe.objects.values("id", "num_documento", "nombre"),
        "receptor_actual": receptor.id if receptor else "",
        "items": items,
        "total": f"{total:.2f}",
    }
    return render(request, "ventas/carrito/carrito.html", context)

@login_required
@require_POST
def carrito_agregar(request):
    """Agrega (o aumenta) un producto para el receptor dado."""
    rid = request.POST.get("receptor_id")
    pid = request.POST.get("producto_id")
    qty = int(request.POST.get("cantidad") or "1")
    precio = request.POST.get("precio")  # opcional (si lo dejas vacío, usamos preunitario)
    iva_on = request.POST.get("iva_on") in ("1","true","True","on")
    desc_pct = request.POST.get("desc_pct") or "0"

    receptor = _ensure_receptor(rid)
    if not receptor:
        return HttpResponseBadRequest("Debe seleccionar un cliente válido.")
    prod = get_object_or_404(Producto, id=pid)

    cart = _get_cart(request)
    bucket = cart.setdefault(str(receptor.id), {})
    row = bucket.setdefault(str(prod.id), {"qty": 0, "precio": None, "iva_on": False, "desc_pct": "0"})
    row["qty"] = max(1, int(row["qty"]) + max(1, qty))
    row["precio"] = precio if precio not in (None, "") else str(prod.preunitario or "0")
    row["iva_on"] = bool(iva_on)
    row["desc_pct"] = desc_pct
    _save_cart(request, cart)

    return JsonResponse({"ok": True})

@login_required
@require_POST
@transaction.atomic
def consolidar_y_redirigir_a_dte(request):
    """
    1. Consulta los detalles de las facturas seleccionadas.
    2. Consolida los productos.
    3. Serializa el resultado en el formato 'prefill' de la sesión.
    4. Redirige a generar_factura_view.
    """
    documentos_ids_str = request.POST.get("documentos_ids")
    tipo_dte_codigo = request.POST.get("tipo_dte")
    
    print(f"TIPO DET codigo en consolidacion >>>>> {tipo_dte_codigo}")

    if not documentos_ids_str:
        return HttpResponseBadRequest(json.dumps({'error': "No se seleccionaron documentos."}), content_type="application/json")

    documentos_ids = [int(id_str) for id_str in documentos_ids_str.split(',') if id_str.isdigit()]

    # 1. Obtener facturas y detalles eficientemente
    facturas_seleccionadas = FacturaElectronica.objects.filter(
        id__in=documentos_ids
    ).prefetch_related(
        'detalles', 
        'detalles__producto',
        'dtereceptor', # Para acceder al receptor sin consulta adicional
    )

    if not facturas_seleccionadas.exists():
        return HttpResponseBadRequest(json.dumps({'error': "No se encontraron facturas válidas."}), content_type="application/json")
    
    # 2. Definir el Receptor y consolidar productos
    primer_factura = facturas_seleccionadas.first()
    receptor_obj = primer_factura.dtereceptor
    
    # Estructura para consolidar productos
    productos_consolidados = {}

    for factura in facturas_seleccionadas:
        for detalle in factura.detalles.all():
            prod = detalle.producto
            prod_id = str(prod.id)
            
            cantidad_actual = float(detalle.cantidad)
            # El precio se toma del detalle, ya que puede ser distinto al precio de lista
            # También lo convertimos a str o float para la sesión
            precio_unitario_float = float(detalle.precio_unitario)
            
            if prod_id in productos_consolidados:
                # Sumar cantidad si el producto ya existe (consolidación)                
                productos_consolidados[prod_id]['cantidad'] += cantidad_actual
                
            else:
                # Inicializar el producto
                productos_consolidados[prod_id] = {
                    "id": prod.id,
                    "cantidad": cantidad_actual,
                    "precio": precio_unitario_float,
                    "nombre": prod.descripcion, 
                    "codigo": prod.codigo,
                    "iva_on": False,
                    "desc_pct": "0", 
                }
                
    # 3. Construir la estructura 'prefill' (simulando el resultado de un carrito facturado)
    prefill_data = {
        'receptor_id': receptor_obj.id,
        'tipo_dte': tipo_dte_codigo,
        'items': list(productos_consolidados.values()),
        'consolidado': True
    }
    # 4. Guardar en la sesión y redirigir
    request.session['facturacion_prefill'] = prefill_data
    request.session.modified = True
    redirect_url = f"{reverse('generar_factura')}?tipo_documento_dte={tipo_dte_codigo}&from_cart=1"
    
    # Redirigir a generar_factura_view con el parámetro from_cart=1
    return JsonResponse({
        "ok": True,
        "redirect_url": redirect_url
    })
    
@login_required
@require_POST
def carrito_actualizar(request):
    """Actualiza cantidad / precio / iva / descuento."""
    rid = request.POST.get("receptor_id")
    pid = request.POST.get("producto_id")
    receptor = _ensure_receptor(rid)
    if not receptor:
        return HttpResponseBadRequest("Cliente inválido")

    cart = _get_cart(request)
    bucket = cart.get(str(receptor.id), {})
    if pid not in bucket:
        return HttpResponseBadRequest("Producto no está en el carrito")

    row = bucket[pid]
    if "cantidad" in request.POST:
        row["qty"] = max(1, int(request.POST.get("cantidad")))
    if "precio" in request.POST and request.POST.get("precio") != "":
        row["precio"] = str(Decimal(str(request.POST.get("precio"))).quantize(Decimal("0.01")))
    if "iva_on" in request.POST:
        row["iva_on"] = request.POST.get("iva_on") in ("1","true","True","on")
    if "desc_pct" in request.POST:
        row["desc_pct"] = str(max(Decimal("0"), min(Decimal("100"), Decimal(str(request.POST.get("desc_pct"))))))
    _save_cart(request, cart)
    return JsonResponse({"ok": True})

@login_required
@require_POST
def carrito_quitar(request):
    rid = request.POST.get("receptor_id")
    pid = request.POST.get("producto_id")
    receptor = _ensure_receptor(rid)
    if not receptor:
        return HttpResponseBadRequest("Cliente inválido")
    cart = _get_cart(request)
    bucket = cart.get(str(receptor.id), {})
    bucket.pop(str(pid), None)
    _save_cart(request, cart)
    return JsonResponse({"ok": True})

@login_required
@require_POST
def carrito_vaciar(request):
    rid = request.POST.get("receptor_id")
    receptor = _ensure_receptor(rid)
    if not receptor:
        return HttpResponseBadRequest("Cliente inválido")
    cart = _get_cart(request)
    cart[str(receptor.id)] = {}
    _save_cart(request, cart)
    return JsonResponse({"ok": True})

@login_required
def carrito_cotizacion(request, receptor_id: int):
    # Renderiza una "prefactura" (no DTE, no firma)
    key = _cart_key(receptor_id)
    carro = request.session.get(key, {})
    if not carro:
        messages.error(request, 'El carrito está vacío.')
        return redirect('carrito_ver', receptor_id=receptor_id)
    receptor = get_object_or_404(Receptor_fe, pk=receptor_id)
    ids = [int(pid) for pid in carro.keys()]
    productos = Producto.objects.filter(id__in=ids)
    items, total = [], Decimal('0.00')
    for p in productos:
        cant = int(carro[str(p.id)])
        precio = (p.preunitario or Decimal('0.00'))
        linea = (precio * Decimal(cant)).quantize(Decimal('0.01'))
        total += linea
        items.append({'producto': p, 'cantidad': cant, 'precio': precio, 'total': linea})
    return render(request, 'ventas/carrito/cotizacion.html', {
        'receptor': receptor,
        'items': items,
        'total': total.quantize(Decimal('0.01'))
    })


@login_required
@require_POST
def carrito_facturar(request):
    rid = request.POST.get("receptor_id")
    receptor = _ensure_receptor(rid)
    if not receptor:
        return HttpResponseBadRequest("Debe seleccionar un cliente válido.")

    cart = _get_cart(request)
    bucket = cart.get(str(receptor.id), {})
    if not bucket:
        return HttpResponseBadRequest("Tu carrito está vacío.")

    items = []
    for pid, row in bucket.items():
        try:
            prod = Producto.objects.get(id=pid)
        except Producto.DoesNotExist:
            continue
        items.append({
            "id": prod.id,
            "codigo": prod.codigo,
            "nombre": prod.descripcion,
            "precio": float(row.get("precio") or prod.preunitario or 0),
            "cantidad": int(row.get("qty") or 1),
            "desc_pct": float(row.get("desc_pct") or 0),
            "iva_on": bool(row.get("iva_on")),
            "stock": int(prod.stock)
        })

    if not items:
        return JsonResponse({"ok": False, "error": "Tu carrito está vacío."}, status=400)


    request.session["facturacion_prefill"] = {
        "receptor_id": receptor.id,
        "items": items,
    }
    request.session.modified = True

    print("ITEMSSS --- ", items)
    # Devolvemos SOLO la URL; no hacemos redirect aquí
    return JsonResponse({
        "ok": True,
        "redirect_url": "/fe/generar/?from_cart=1"
    })


@login_required
def carrito_add_go(request):
    """GET /carrito/add-go/?rid=&pid=&qty=  -> agrega y redirige al carrito."""
    rid = request.GET.get("rid")
    pid = request.GET.get("pid")
    qty = int(request.GET.get("qty") or "1")
    receptor = _ensure_receptor(rid)
    if not receptor:
        return HttpResponseBadRequest("Cliente inválido")
    prod = get_object_or_404(Producto, id=pid)
    cart = _get_cart(request)
    bucket = cart.setdefault(str(receptor.id), {})
    row = bucket.setdefault(str(prod.id), {"qty": 0, "precio": str(prod.preunitario or "0"), "iva_on": False, "desc_pct": "0"})
    row["qty"] = max(1, int(row["qty"]) + max(1, qty))
    _save_cart(request, cart)
    from django.urls import reverse
    return redirect(f'{reverse("carrito_ver")}?receptor_id={receptor.id}')

# ---------- LISTA/DETALLE VENTAS ---------
@login_required
def ventas_list(request):
    qs = FacturaElectronica.objects.select_related(
        'dtereceptor__municipio__departamento',
        'tipo_dte'
    ).all().order_by('-fecha_emision', '-id')

    # --- FILTROS ---
    tipo = request.GET.get('tipo', '')
    cliente = request.GET.get('cliente', '')
    producto = request.GET.get('producto', '')
    usuario = request.GET.get('usuario', '')
    fecha_ini = request.GET.get('fecha_ini', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    estado = request.GET.get('estado', '')
    departamento = request.GET.get('departamento', '')
    municipio = request.GET.get('municipio', '')
    monto_min = request.GET.get('monto_min', '')
    monto_max = request.GET.get('monto_max', '')

    if tipo:
        qs = qs.filter(tipo_dte__codigo=tipo)
    if cliente:
        qs = qs.filter(dtereceptor__id=cliente)
    if fecha_ini:
        qs = qs.filter(fecha_emision__gte=fecha_ini)
    if fecha_fin:
        qs = qs.filter(fecha_emision__lte=fecha_fin)
    if estado == "firmado":
        qs = qs.filter(firmado=True, recibido_mh=False)
    elif estado == "recibido":
        qs = qs.filter(recibido_mh=True)
    elif estado == "borrador":
        qs = qs.filter(firmado=False)

    if departamento:
        qs = qs.filter(dtereceptor__municipio__departamento__id=departamento)
    if municipio:
        qs = qs.filter(dtereceptor__municipio__id=municipio)

    if monto_min:
        qs = qs.filter(total_pagar__gte=monto_min)
    if monto_max:
        qs = qs.filter(total_pagar__lte=monto_max)

    if producto:
        qs = qs.filter(detalles__producto__id=producto).distinct()

    if usuario:
        qs = qs.filter(json_original__user_id=usuario)  # ⚠️ ajusta si tienes un campo real "usuario"
        # si guardas el usuario en otro campo de FacturaElectronica cámbialo aquí

    # --- PAGINACIÓN ---
    paginator = Paginator(qs, 20)
    page = request.GET.get('page')
    ventas_page = paginator.get_page(page)

    # --- Limpieza del querystring (para paginación) ---
    params = request.GET.copy()
    if "page" in params:
        params.pop("page")
    querystring = params.urlencode()

    # --- RESUMEN ---
    resumen = qs.aggregate(
        total_ventas=Sum('total_pagar'),
        cantidad=Count('id')
    )

    context = {
        'ventas': ventas_page,
        'resumen': resumen,
        'tipo': tipo,
        'cliente': cliente,
        'producto': producto,
        'usuario': usuario,
        'fecha_ini': fecha_ini,
        'fecha_fin': fecha_fin,
        'estado': estado,
        'departamento': departamento,
        'municipio': municipio,
        'monto_min': monto_min,
        'monto_max': monto_max,

        # listas para selects
        'clientes': Receptor_fe.objects.all().order_by('nombre'),
        'productos': Producto.objects.all().order_by('descripcion'),
        'usuarios': User.objects.all().order_by('username'),
        **request.GET.dict(),
        'querystring': querystring,
    }
    return render(request, 'ventas/lista.html', context)

@login_required
def venta_detalle(request, factura_id: int):
    # Reusa tu template existente del detalle
    return redirect('detalle_factura', factura_id=factura_id)

# ---------- DEVOLUCIONES DE VENTA ----------
@login_required
def devoluciones_list(request):
    devs = DevolucionVenta.objects.select_related().all().order_by('-fecha')
    return render(request, 'ventas/devoluciones/lista.html', {'devoluciones': devs})

@transaction.atomic
def devolucion_crear(request, factura_id: int):
    factura = get_object_or_404(FacturaElectronica.objects.prefetch_related('detalles__producto'), pk=factura_id)
    if request.method == 'POST':
        motivo = request.POST.get('motivo') or 'Devolución de venta'
        usuario = request.user.username if request.user.is_authenticated else None
        devolucion = DevolucionVenta.objects.create(
            num_factura=factura.numero_control or f'FAC-{factura.id}',
            motivo=motivo,
            estado='Aprobada',
            usuario=usuario
        )
        # Recoger cantidades a devolver por línea
        for det in factura.detalles.all():
            cant = int(request.POST.get(f'cant_{det.id}', 0) or 0)
            if cant <= 0:
                continue
            # Crear detalle de devolución
            DetalleDevolucionVenta.objects.create(
                devolucion=devolucion,
                producto=det.producto,
                cantidad=cant,
                motivo_detalle=motivo
            )
            # Movimiento de inventario (Entrada) — EL STOCK SE AJUSTA EN EL SIGNAL
            almacen = det.producto.almacenes.first() if det.producto.almacenes.exists() else Almacen.objects.first()
            MovimientoInventario.objects.create(
                producto=det.producto,
                almacen=almacen,
                tipo='Entrada',
                cantidad=cant,
                referencia=f'Devolución Factura {factura.codigo_generacion or factura.id}'
            )
        messages.success(request, 'Devolución registrada.')
        return redirect('devoluciones_list')

    return render(request, 'ventas/devoluciones/form.html', {'factura': factura})



# FE/views.py  (o donde tengas las vistas de facturación)
from django.http import JsonResponse
from django.db.models import Q
from INVENTARIO.models import Producto

@login_required
def api_productos(request):
    """
    Devuelve productos en JSON para el catálogo del modal.
    Parámetros:
      - q: texto de búsqueda (código o descripción)
      - page: página (1 por defecto)
      - page_size: tamaño de página (24 por defecto)
    """
    q = (request.GET.get('q') or '').strip()
    try:
        page = max(1, int(request.GET.get('page', 1)))
    except ValueError:
        page = 1
    try:
        page_size = max(1, min(96, int(request.GET.get('page_size', 24))))
    except ValueError:
        page_size = 24

    qs = Producto.objects.all().order_by('descripcion')
    if q:
        qs = qs.filter(Q(descripcion__icontains=q) | Q(codigo__icontains=q))

    total = qs.count()
    start = (page - 1) * page_size
    end = start + page_size

    results = []
    for p in qs[start:end]:
        results.append({
            'id': p.id,
            'codigo': p.codigo,
            'nombre': p.descripcion,
            'precio': str(p.precio_venta),
            'stock': p.stock,
            'imagen': (p.imagen.url if getattr(p, 'imagen', None) else ''),
            'con_iva': bool(getattr(p, 'precio_iva', False)),
        })

    return JsonResponse({
        'results': results,
        'page': page,
        'has_next': end < total,
        'total': total,
    })

#################################
# VISTA NUEVA PARA GENERAR FACTURAS

@login_required
def generar_factura(request):

    print ("Generando factura...", request)

    # variables globales arays
    productos_ids_r = []
    productos_cant_r = []
    documentos_relacionador = []
    descuentos_r = []
    precios_r = []

    #contador de intentos en session
    if 'intentos' not in request.session:
        request.session['intentos'] = 0
    
    # metodo get para cargar la vista
    if request.method == 'GET':
        # cargar datos de productos
        productos = Producto.objects.all().order_by('descripcion')
        # cargar datos de clientes
        clientes = Receptor_fe.objects.all().order_by('nombre')
        # cargar datos de tipos de documento
        tipos_documento = TiposDocIDReceptor.objects.all().order_by('descripcion')
        # cargar datos de municipios
        municipios = Municipio.objects.select_related('departamento').order_by('descripcion')
        # cargar datos de unidades de medida
        unidades_medida = TipoUnidadMedida.objects.all().order_by('descripcion')
        # cargar datos de tributos
        tributos = Tributo.objects.all().order_by('descripcion')
        # cargar datos de tipos de documento (DTE)
        tipos_dte = Tipo_dte.objects.all().order_by('descripcion')
        # cargar datos de modelos de facturacion
        modelos_facturacion = Modelofacturacion.objects.all().order_by('descripcion')
        # cargar datos de tipos de transmision
        tipos_transmision = TipoTransmision.objects.all().order_by('descripcion')

        tipos_dte = Tipo_dte.obj
        emisor_obj = None
        if emisor_obj:
            new_num_control = NumeroControl.preview_numero_control(emisor_obj)



    return render(request, 'ventas/generar_factura.html')


# REPORTES

# REPORTES PARA CONTABILIDAD

## HELPER
def _qs_without_page(request):
    params = request.GET.copy()
    params.pop('page', None)
    return params.urlencode()

##### REPORTES
@login_required
def reporte_contabilidad_view(request):
    """
    REPORTE PARA CONTABILIDAD-INCOE
    - fecha_emision
    - tipo (01,03,05,06)
    - numero_control
    - sello_recepcion
    - codigo_generacion
    - si es 03: NRC, NIT, nombre cliente
    - total grabado + IVA
    - total IVA
    - total grabado sin IVA
    Con filtros, paginación y exportación a Excel.
    """
    # --------- Filtros ---------
    hoy = datetime.now().date()
    fini = request.GET.get("fini") or hoy.replace(day=1).isoformat()
    ffin = request.GET.get("ffin") or hoy.isoformat()
    tipos = request.GET.getlist("tipo")  # lista: ["01","03",...]
    buscar = (request.GET.get("q") or "").strip()
    page = int(request.GET.get("page") or 1)
    per_page = int(request.GET.get("per_page") or 25)
    export = request.GET.get("export") == "excel"

    tipos_validos = {"01", "03", "05", "06"}
    if not tipos:
        tipos = ["01", "03", "05", "06"]
    else:
        tipos = [t for t in tipos if t in tipos_validos]
        if not tipos:
            tipos = ["01", "03", "05", "06"]

    # --------- QuerySet base ---------
    qs = (
        FacturaElectronica.objects
        .select_related("tipo_dte", "dtereceptor")
        .filter(fecha_emision__range=[fini, ffin],
                tipo_dte__codigo__in=tipos)
        .order_by("-fecha_emision", "-id")
    )

    if buscar:
        qs = qs.filter(
            Q(numero_control__icontains=buscar) |
            Q(sello_recepcion__icontains=buscar) |
            Q(codigo_generacion__icontains=buscar)
        )


    # --------- Anotaciones (2 pasos) ---------
    cero = Value(Decimal("0.00"), output_field=DecimalField(max_digits=15, decimal_places=2))
    tasa_iva = Value(Decimal("0.13"), output_field=DecimalField(max_digits=5, decimal_places=2))

    # 1) Sumar desde detalles
    qs = qs.annotate(
        sum_det_iva_item=Coalesce(Sum("detalles__iva_item"), cero),
        sum_det_gravadas=Coalesce(Sum("detalles__ventas_gravadas"), cero),
    )

    # 2) Campos calculados usando las sumas anteriores
    qs = qs.annotate(
        # Total gravadas: usa header, si viene NULL/0 usa la suma del detalle
        a_total_gravadas=Case(
            When(Q(total_gravadas__isnull=True) | Q(total_gravadas=0), then=F("sum_det_gravadas")),
            default=Coalesce(F("total_gravadas"), cero),
            output_field=DecimalField(max_digits=15, decimal_places=2),
        ),
        # IVA estimado desde detalle: prioridad SUM(iva_item) > (SUM(gravadas) * 0.13)
        a_total_iva=Case(
            When(Q(total_iva__isnull=True) | Q(total_iva=0),
                then=Case(
                    When(sum_det_iva_item__gt=Decimal("0.00"), then=F("sum_det_iva_item")),
                    default=ExpressionWrapper(F("sum_det_gravadas") * tasa_iva,
                                            output_field=DecimalField(max_digits=15, decimal_places=2)),
                )),
            default=Coalesce(F("total_iva"), cero),
            output_field=DecimalField(max_digits=15, decimal_places=2),
        ),
    )

    # Total grabado + IVA
    qs = qs.annotate(
        a_total_grabado_mas_iva=ExpressionWrapper(
            F("a_total_gravadas"),
            output_field=DecimalField(max_digits=15, decimal_places=2),
        ),
    )

    # Total gravado sin IVA (para reporte)
    qs = qs.annotate(
        a_total_grabado_sin_iva=ExpressionWrapper(
            F("a_total_gravadas") - F("a_total_iva"),
            output_field=DecimalField(max_digits=15, decimal_places=2),
        ),
    )

    # --------- Exportar a Excel ---------
    if export:
        return _reporte_contabilidad_excel(qs, fini, ffin, tipos, buscar)

    # --------- Paginación ---------
    paginator = Paginator(qs, per_page)
    page_obj = paginator.get_page(page)

    ctx = {
        "fini": fini,
        "ffin": ffin,
        "tipos": tipos,
        "buscar": buscar,
        "page_obj": page_obj,
        "qs_no_page": _qs_without_page(request),
        "per_page": per_page,
        "tipos_opciones": [("01", "01 - Factura Electrónica"),
                           ("03", "03 - Comprobante Crédito Fiscal"),
                           ("05", "05 - Nota de Crédito"),
                           ("06", "06 - Nota de Débito")],
    }
    return render(request, "reportes/reporte_contabilidad.html", ctx)

# ---------------- Excel ----------------
def _reporte_contabilidad_excel(qs, fini, ffin, tipos, buscar):

    """
    Exporta el queryset filtrado a Excel.
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment, Font
    except ImportError:
        return HttpResponse("Falta dependencia 'openpyxl' para exportar a Excel.", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contabilidad"

    title = f"REPORTE CONTABILIDAD INCOE ({fini} a {ffin})"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(horizontal="center")

    headers = [
        "Fecha emisión",
        "Tipo",
        "Número de control",
        "Sello recepción",
        "Código generación",
        "NRC (si 03)",
        "NIT (si 03)",
        "Cliente (si 03)",
        "Total grabado + IVA",
        "Total IVA",
        "Total grabado sin IVA",
        "Recibido MH",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True)

    row = 4
    for f in qs:
        es_03 = (str(getattr(f.tipo_dte, "codigo", "")) == "03")
        nrc = f.dtereceptor.nrc if (es_03 and f.dtereceptor) else ""
        nit = f.dtereceptor.num_documento if (es_03 and f.dtereceptor) else ""
        cliente = f.dtereceptor.nombre if (es_03 and f.dtereceptor) else ""

        ws.cell(row=row, column=1, value=f.fecha_emision.isoformat() if f.fecha_emision else "")
        ws.cell(row=row, column=2, value=getattr(f.tipo_dte, "codigo", ""))
        ws.cell(row=row, column=3, value=f.numero_control)
        ws.cell(row=row, column=4, value=f.sello_recepcion or "")
        ws.cell(row=row, column=5, value=str(f.codigo_generacion))
        ws.cell(row=row, column=6, value=nrc)
        ws.cell(row=row, column=7, value=nit)
        ws.cell(row=row, column=8, value=cliente)
        ws.cell(row=row, column=9, value=float(f.a_total_grabado_mas_iva or 0))
        ws.cell(row=row, column=10, value=float(f.a_total_iva or 0))
        ws.cell(row=row, column=11, value=float(f.a_total_gravadas or 0))
        ws.cell(row=row, column=12, value="Sí" if f.recibido_mh else "No")
        row += 1

    # Auto ancho simple
    for col in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"reporte_contabilidad_{fini}_{ffin}.xlsx"
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp

#################################

@login_required
def reporte_facturacion_view(request):
    """
    REPORTE PARA FACTURACION-INCOE

    Encabezado por factura:
      - fecha_emision
      - numero_control
      - sello_recepcion
      - codigo_generacion

    Ítems por factura:
      - código, descripción, cantidad, precio_unitario, total_pagar_item
        (total_pagar_item = cantidad * precio_unitario + iva_item; si iva_item es None → 0)
    """

    # --------- Filtros ---------
    hoy = date.today()
    fini = request.GET.get("fini") or hoy.replace(day=1).isoformat()
    ffin = request.GET.get("ffin") or hoy.isoformat()
    buscar = (request.GET.get("q") or "").strip()
    tipos = request.GET.getlist("tipo")
    page = int(request.GET.get("page") or 1)
    per_page = int(request.GET.get("per_page") or 25)
    export = (request.GET.get("export") == "excel")

    tipos_validos = {"01", "03", "05", "06"}
    if not tipos:
        tipos = ["01", "03", "05", "06"]
    else:
        tipos = [t for t in tipos if t in tipos_validos] or ["01", "03", "05", "06"]

    # --------- Query base Facturas ---------
    qs = (
        FacturaElectronica.objects
        .select_related("tipo_dte")
        .filter(
            fecha_emision__range=[fini, ffin],
            tipo_dte__codigo__in=tipos
        )
        .order_by("-fecha_emision", "-id")
    )

    if buscar:
        qs = qs.filter(
            Q(numero_control__icontains=buscar) |
            Q(sello_recepcion__icontains=buscar) |
            Q(codigo_generacion__icontains=buscar)
        )

    # --------- Prefetch de Detalles con anotaciones ---------
    cero = Value(Decimal("0.00"), output_field=DecimalField(max_digits=16, decimal_places=2))
    detalles_qs = (
        DetalleFactura.objects
        .select_related("producto", "unidad_medida")
        .annotate(
            a_cantidad=Coalesce(F("cantidad"), Value(0)),
            a_precio=Coalesce(F("precio_unitario"), cero),
            a_iva=Coalesce(F("iva_item"), cero),
        )
        .annotate(
            total_pagar_item=ExpressionWrapper(
                F("a_cantidad") * F("a_precio") + F("a_iva"),
                output_field=DecimalField(max_digits=16, decimal_places=2),
            )
        )
        .order_by("id")
    )

    qs = qs.prefetch_related(
        Prefetch("detalles", queryset=detalles_qs, to_attr="detalles_anno")
    )

    # --------- Exportar a Excel ---------
    if export:
        return _reporte_facturacion_excel(qs, fini, ffin, tipos, buscar)

    # --------- Paginación ---------
    paginator = Paginator(qs, per_page)
    page_obj = paginator.get_page(page)

    ctx = {
        "fini": fini,
        "ffin": ffin,
        "tipos": tipos,
        "buscar": buscar,
        "page_obj": page_obj,
        "qs_no_page": _qs_without_page(request),
        "per_page": per_page,
        "tipos_opciones": [
            ("01", "01 - Factura Electrónica"),
            ("03", "03 - Comprobante Crédito Fiscal"),
            ("05", "05 - Nota de Crédito"),
            ("06", "06 - Nota de Débito"),
        ],
    }
    return render(request, "reportes/reporte_facturacion.html", ctx)


def _reporte_facturacion_excel(qs, fini, ffin, tipos=None, buscar=""):
    tipos = tipos or []
    """
    Exporta a Excel:
      Encabezado factura + filas por cada ítem.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturación INCOE"

    headers = [
        "Fecha emisión",
        "Tipo DTE",
        "Número control",
        "Sello recepción",
        "Código generación",
        "Código ítem",
        "Descripción ítem",
        "Cantidad",
        "Precio unitario",
        "IVA ítem",
        "Total ítem",
    ]
    ws.append(headers)

    for f in qs:
        for it in getattr(f, "detalles_anno", []):
            ws.append([
                smart_str(f.fecha_emision or ""),
                smart_str(getattr(f.tipo_dte, "codigo", "")),
                smart_str(f.numero_control or ""),
                smart_str(f.sello_recepcion or ""),
                smart_str(f.codigo_generacion or ""),
                smart_str(getattr(it.producto, "codigo", "") or ""),
                smart_str(getattr(it.producto, "descripcion", "") or str(getattr(it.producto, "id", "") or "")),
                float(it.cantidad or 0),
                float(it.precio_unitario or 0),
                float(it.iva_item or 0),
                float(getattr(it, "total_pagar_item", 0) or 0),
            ])

    # Ancho de columnas (opcional)
    widths = [14, 8, 24, 22, 40, 14, 36, 10, 14, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    fname = f"reporte_facturacion_{fini}_a_{ffin}.xlsx"
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    wb.save(resp)
    return resp


